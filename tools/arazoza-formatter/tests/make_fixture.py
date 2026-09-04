"""Fixture builders for the arazoza-formatter tests.

Two ways to get a *raw* Arazoza worksheet (the shape the tool expects as input):

  build_synthetic_raw(path)         a hand-built workbook that exercises every rule and every
                                    edge the spec calls out (red section markers, ambiguous
                                    headers, depth items, package tokens, conflicts, ...).
  deformat_finished(src, dst)       reverse a *finished* Arazoza worksheet (we have several
                                    real ones) back into the raw layout, so the tool's output
                                    can be compared cell-for-cell with what a human produced.

Also runnable from the terminal to eyeball what the tool expects:
    .venv/bin/python tests/make_fixture.py /tmp/raw.xlsx
"""

from __future__ import annotations

import sys
from pathlib import Path

import datetime as dt

import openpyxl
from openpyxl.styles import Font

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

BLACK_THEME = Font(bold=True, color=None)
RED = Font(bold=True, color="FFFF0000")

HEADERS = ["Description", "Size", "Package", "Notes", "Qty 1", "UOM", "Qty 2", "UOM2", "Column1"]
HEADER_ROW = 15
FIRST_COL = 2   # B


def _put(ws, row: int, values: dict, font: Font | None = None):
    """values: {'B': ..., 'C': ...}."""
    for col_letter, v in values.items():
        c = ws[f"{col_letter}{row}"]
        c.value = v
        if font is not None:
            c.font = font


# The synthetic raw sheet, row by row. Each entry: (row, {col: value}, font-or-None).
# The expectations for every one of these rows live in test_formatter.py.
SYNTHETIC_ROWS = [
    (16, {"B": "Landscape"}, RED),                                        # coloured section marker: skipped
    (18, {"B": "Trees"}, BLACK_THEME),                                    # -> 1 - Trees & Palms
    (19, {"C": "Quercus virginiana / Live Oak", "E": "45 G, 12'-14' Ht. x 6'-7' Sp., 3\" Cal",
          "F": 68, "G": "EA", "H": 1900, "I": "SF"}, None),
    (20, {"C": "Sabal palmetto / Cabbage Palm", "E": "14\" Cal, 10'-16' CT, FG", "F": 83, "G": "ea."}, None),
    (21, {"C": "Bauhinia x blakeana / Hong Kong Orchid Tree", "E": "100 G Cont., 14' Ht.", "F": 18, "G": "Each"}, None),
    (22, {"B": "Palms"}, BLACK_THEME),                                    # -> 1 - Trees & Palms
    (23, {"C": "Roystonea regia / Royal Palm", "E": "16' Ht. 12\" Cal. B&B", "F": 4, "G": "Count"}, None),
    (24, {"C": "Phoenix dactylifera / Date Palm", "D": "Container", "E": "FG / B & B, 20' CT", "F": 2, "G": "EA"}, None),
    (25, {"B": "Shrubs & Groundcover"}, BLACK_THEME),                     # ambiguous -> RED
    (26, {"C": "Clusia guttifera / Small Leaf Clusia", "E": "3 G, 36\" Ht. x 24\" Sp.", "F": 283, "G": "EA"}, None),
    (27, {"C": "Ficus microcarpa Green Island - 3 G", "E": "3 G", "F": 300, "G": "EA"}, None),   # size already in text
    (28, {"B": "Mulch"}, BLACK_THEME),                                    # -> 31 - Mulch (SY)
    (29, {"C": "Mulch", "E": "3\" Depth", "F": 36327, "G": "SF", "H": 336, "I": "CY"}, None),
    (30, {"C": "Assumed @ 4' Dia of Mulch Ring", "F": 199, "G": "EA"}, None),
    (31, {"C": "Pine Bark Mulch", "E": "3\"", "F": 100, "G": "sf"}, None),   # generic size, no 'depth'
    (32, {"B": "Soil"}, BLACK_THEME),                                     # -> 4 - Soil
    (33, {"C": "Planting Soil Mix 12\" Depth", "F": 14440, "G": "SF"}, None),
    (34, {"C": "Amended bed mix 6\" depth", "F": 500, "G": "SF"}, None),     # soil by GROUP, not by name
    (35, {"C": "Topsoil for Grass Areas", "E": "2\" Depth", "F": 1509, "G": "SF"}, None),
    (36, {"B": "Bed Prep"}, BLACK_THEME),                                 # -> 154 - Bed Preparation
    (37, {"C": "Bed Edging", "F": 2946, "G": "LF"}, None),
    (38, {"C": "Weed Barrier 3\" depth", "F": 10, "G": "L.F."}, None),      # depth, but not soil/mulch
    (39, {"B": "Sod"}, BLACK_THEME),                                      # -> 15 - Sod (SF)
    (40, {"C": "Assume Sod", "F": 628362, "G": "S.F.", "I": "CY"}, None),
    (41, {"B": "Irrigation Sleeves"}, BLACK_THEME),                       # no match -> RED
    (42, {"C": "4\" PVC Sleeve", "E": "Sch 40", "F": 120, "G": "lf"}, None),
    (43, {"B": "1 - Trees & Palms"}, BLACK_THEME),                        # already in final form
    (44, {"C": "Live Oak", "E": "65 gal", "F": 1, "G": "CY"}, None),         # CY is not converted
    # B and C both filled -> RED, untouched: not relabelled, not merged, no UOM, and no Package
    # even though its Size carries a package token
    (45, {"B": "Root Barrier", "C": "Conflict item, cont", "G": "EA"}, BLACK_THEME),
    (46, {"B": "PLANTING PLAN - SITE"}, BLACK_THEME),                     # no match -> RED
    (47, {"C": "Depth of mulch unknown", "F": 1, "G": "EA"}, None),          # 'depth' with no measurement
    (48, {"C": "Numeric Note Item", "E": 3, "F": 1, "G": "EA"}, None),       # numeric Notes value
    (49, {"B": "Shade Trees"}, BLACK_THEME),                              # adjective + noun -> Trees & Palms
    (50, {"C": "Ulmus parvifolia / Drake Elm", "E": "3\" cal, cont", "F": 5, "G": "EA", "J": 5}, None),
    (51, {"B": "Mulch", "E": "3\" depth cont"}, BLACK_THEME),              # header WITH a note: note stays put
    (52, {"C": "Pine Straw Mulch", "E": "bulk cont., 3\" depth", "F": 50, "G": "SF"}, None),   # depth beats package (5 before 7)
    (53, {"B": "Decorative Stone"}, BLACK_THEME),                         # no match -> RED; ends the Mulch group
    (54, {"C": "Crushed Granite", "E": "4\" depth, cont", "F": 200, "G": "SF"}, None),        # NOT soil/mulch: no depth rule
    (55, {"B": "IRRIGATION", "E": "per IR-1"}, RED),                      # coloured marker with a note: untouched
    (56, {"C": "2\" PVC Mainline", "E": "Sch 40, 18\" depth cover", "F": 300, "G": "LF"}, None),   # irrigation: no depth rule
    (57, {"C": "Date Note Item", "E": dt.datetime(2023, 3, 4), "F": 1, "G": "EA"}, None),   # Excel turned '3-4' into a date
    (58, {"C": "Mulch 2 1/2\" depth", "F": 40, "G": "SF"}, None),           # fraction depth
    # Description-only rows that carry a Qty are items somebody already typed into the right
    # column — never group headers, so their names must survive step 1 untouched.
    (59, {"B": "Pine Straw Mulch 3\" depth", "F": 10, "G": "SF"}, None),    # + 'pine straw' IS mulch
    (60, {"B": "Existing Tree to Remain", "F": 3, "G": "EA"}, None),
]


def build_synthetic_raw(path: str | Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Project Totals"
    # metadata block like the real template — the header-row finder must skip past it
    ws["A1"], ws["B1"] = "Company Name: ", "Arazoza Brothers"
    ws["A9"], ws["B9"] = "Project Name: ", "Synthetic Fixture"
    ws["B14"], ws["C14"], ws["F14"] = "product_name", "pkg_size_combined", "quantity"
    for i, h in enumerate(HEADERS):
        c = ws.cell(HEADER_ROW, FIRST_COL + i, h)
        c.font = Font(bold=True)
    for row, values, font in SYNTHETIC_ROWS:
        _put(ws, row, values, font)
    # a decoy hidden sheet with the same headers and no data (the LH template has four)
    decoy = wb.create_sheet("Project Break-Out 1")
    for i, h in enumerate(HEADERS):
        decoy.cell(8, FIRST_COL + i, h)
    decoy.sheet_state = "hidden"
    path = Path(path)
    wb.save(path)
    return path


def build_already_formatted(path: str | Path) -> Path:
    """Items already in Description, sizes in Size, nothing in Notes — the tool must refuse."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Totals"
    for i, h in enumerate(HEADERS):
        ws.cell(HEADER_ROW, FIRST_COL + i, h)
    _put(ws, 16, {"B": "1 - Trees & Palms"}, BLACK_THEME)
    _put(ws, 17, {"B": "Quercus virginiana / Live Oak - 45 G", "C": "45 G", "D": "Container", "F": 68, "G": "Unit"})
    _put(ws, 18, {"B": "Sabal palmetto / Cabbage Palm - 14\" Cal", "C": "14\" Cal", "F": 83, "G": "Unit"})
    path = Path(path)
    wb.save(path)
    return path


# --- reversing a finished worksheet ------------------------------------------------------

# How a human might have labelled these groups on the raw sheet. Rotated per occurrence so a
# single test exercises one-word, two-word and already-final spellings.
RAW_NAMES = {
    "1 - Trees & Palms": ["Trees", "Palms", "Trees & Palms", "Palm Trees", "1 - Trees & Palms"],
    "2 - Shrubs": ["Shrubs", "Shrub", "2 - Shrubs"],
    "9 - Groundcover": ["Groundcovers", "Ground Cover", "Groundcover"],
    "31 - Mulch (SY)": ["Mulch", "Mulching"],
    "154 - Bed Preparation": ["Bed Prep", "Bed Preparation"],
    "15 - Sod (SF)": ["Sod", "Sod (SF)"],
    "4 - Soil": ["Soil", "Soils", "Planting Soil"],
    "26 - Root Barrier": ["Root Barrier", "Root Barriers"],
}
RAW_UOM = {"Unit": ["EA", "ea.", "Each", "count", "EA"], "Square Feet": ["SF", "sf", "S.F."], "Linear Feet": ["LF", "lf", "L.F."]}


def deformat_finished(src: str | Path, dst: str | Path, sheet: str = "Project Totals") -> dict:
    """Turn a finished Arazoza worksheet back into the raw layout the tool expects.

    For each row under the header:
      - a header row (Description filled, no Qty):  label -> one of RAW_NAMES
      - an item row:  Description "desc - size" -> Size=desc, Notes=size, Description empty;
                      Package cleared; UOM 'Unit'/'Square Feet'/'Linear Feet' -> raw spellings.
    Returns {'expected': {row: {'B':..,'C':..,'D':..,'G':..}}, 'skipped': [(row, why)]} — the
    finished values the tool should reproduce (D only asserted on 'Depth' rows: the human-made
    sheets say 'Container' for gallon sizes, a rule the spec does not include). Rows the human
    left internally inconsistent can't be reversed and are skipped: a Size that isn't the
    description's ' - ' suffix ("1Gal." vs "1 Gal."), or a depth item whose Size/Package were
    never filled in."""
    import formatter

    wb = openpyxl.load_workbook(src, keep_vba=str(src).lower().endswith(".xlsm"))
    ws = wb[sheet]
    layout = formatter.find_layout(wb, sheet)
    desc_c, size_c, pkg_c, notes_c, uom_c, qty_c = (layout.cols[k] for k in ("description", "size", "package", "notes", "uom", "qty"))
    counters: dict = {}
    expected: dict = {}
    skipped: list = []
    for r in layout.rows():
        desc = ws.cell(r, desc_c)
        size = ws.cell(r, size_c)
        pkg = ws.cell(r, pkg_c)
        uom = ws.cell(r, uom_c)
        qty = ws.cell(r, qty_c)
        d = formatter._txt(desc.value)
        if not d:
            continue
        expected[r] = {"B": d, "C": formatter._txt(size.value), "D": formatter._txt(pkg.value), "G": formatter._txt(uom.value)}
        is_header = formatter._empty(qty.value) and formatter._empty(size.value)
        if is_header:
            if d in RAW_NAMES and formatter.is_black_font(desc):
                i = counters.get(d, 0)
                counters[d] = i + 1
                desc.value = RAW_NAMES[d][i % len(RAW_NAMES[d])]
            continue
        s = formatter._txt(size.value)
        suffix = f" - {s}"
        if s and not d.endswith(suffix) and formatter._txt(pkg.value) != "Depth":
            # (a depth row's Size is the depth, not the suffix — that one the tool reproduces)
            skipped.append((r, f"Size {s!r} is not the description's suffix"))
            expected.pop(r)
        elif not s and formatter.DEPTH_RE.search(d) and formatter.SOIL_MULCH_RE.search(d):
            skipped.append((r, "depth item whose Size/Package the human never filled in"))
            expected.pop(r)
        item = d[: -len(suffix)] if s and d.endswith(suffix) else d
        size.value = item
        desc.value = None
        pkg.value = None
        if s and d.endswith(suffix):
            ws.cell(r, notes_c).value = s
            size_from_notes = True
        else:
            size_from_notes = False
        if not size_from_notes:
            # the finished Size on depth rows ('3"') was derived from the description, not
            # from Notes — so the raw sheet had nothing in Notes there.
            pass
        u = formatter._txt(uom.value)
        if u in RAW_UOM:
            i = counters.get(("uom", u), 0)
            counters[("uom", u)] = i + 1
            uom.value = RAW_UOM[u][i % len(RAW_UOM[u])]
    wb.save(dst)
    return {"expected": expected, "skipped": skipped}


if __name__ == "__main__":
    out = build_synthetic_raw(sys.argv[1] if len(sys.argv) > 1 else "arazoza-raw-fixture.xlsx")
    print(f"wrote {out}")
