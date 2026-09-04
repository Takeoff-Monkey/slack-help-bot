"""Tests for the arazoza-formatter tool.

Run from the tool directory with its own venv:
    .venv/bin/python -m unittest discover -s tests -v

The synthetic fixture (tests/make_fixture.py) pins down every rule in TODO #6 and the
order they run in. The real-sample test reverses one of the finished Arazoza worksheets in
../../../Sample_test_sheets into the raw layout and checks the tool rebuilds what a human
produced; it is skipped when that folder isn't on this machine.
"""

from __future__ import annotations

import json
import os
import subprocess
import sys
import tempfile
import unittest
import zipfile
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.colors import Color

HERE = Path(__file__).resolve().parent
TOOL_DIR = HERE.parent
sys.path.insert(0, str(TOOL_DIR))
sys.path.insert(0, str(HERE))

import formatter            # noqa: E402
import make_fixture         # noqa: E402

_DEFAULT_SAMPLES = TOOL_DIR.parents[2] / "Sample_test_sheets" if len(TOOL_DIR.parents) > 2 else Path("/nonexistent")
SAMPLE_DIR = Path(os.environ.get("ARAZOZA_SAMPLE_DIR", _DEFAULT_SAMPLES))
FINISHED_SAMPLE = SAMPLE_DIR / "Arazoza - Esplanade at Tradition Ph 4 - Worksheet 2023-10-09.xlsm"

# The one genuine raw OST export we have, with the worksheet a human finished from it. Looked
# for in tests/samples/ first, then the repo's docs/ folder where they were first dropped.
REPO = TOOL_DIR.parents[1]
_VELA_DIRS = (HERE / "samples", REPO / "docs")
VELA_RAW = next((d / "Vela Cove OST Output.xlsm" for d in _VELA_DIRS if (d / "Vela Cove OST Output.xlsm").exists()), None)
VELA_FINISHED = next((d / "Arazoza - Vela Cove - Worksheet 2026-09-01.xlsm" for d in _VELA_DIRS
                      if (d / "Arazoza - Vela Cove - Worksheet 2026-09-01.xlsm").exists()), None)


def _red(cell) -> bool:
    return cell.fill is not None and cell.fill.fill_type == "solid" and str(cell.fill.start_color.rgb).upper().endswith("FF0000")


class GroupMatchingTests(unittest.TestCase):
    def setUp(self):
        self.groupings = formatter.load_groupings()

    def test_groupings_file_loads_every_line(self):
        labels = [g.label for g in self.groupings]
        self.assertIn("1 - Trees & Palms", labels)
        self.assertIn("29 - Rock & Boulders", labels)
        self.assertEqual(len(labels), 17)

    def test_matches(self):
        cases = {
            "Trees": "1 - Trees & Palms",
            "TREES": "1 - Trees & Palms",
            "Palms": "1 - Trees & Palms",
            "Palm Trees": "1 - Trees & Palms",
            "Shade Trees": "1 - Trees & Palms",
            "Trees & Palms": "1 - Trees & Palms",
            "1 - Trees & Palms": "1 - Trees & Palms",
            "01 - Trees": "1 - Trees & Palms",
            "Shrubs": "2 - Shrubs",
            "Shrub": "2 - Shrubs",
            "Accent Shrubs": "2 - Shrubs",
            "Groundcovers": "9 - Groundcover",
            "Ground Cover": "9 - Groundcover",
            "Mulch": "31 - Mulch (SY)",
            "Mulching": "31 - Mulch (SY)",
            "Mulch (SY)": "31 - Mulch (SY)",
            "Soil": "4 - Soil",
            "Soils": "4 - Soil",
            "Planting Soil": "4 - Soil",
            "Topsoil": "4 - Soil",
            "Sod": "15 - Sod (SF)",
            "Sod (SF)": "15 - Sod (SF)",
            "Turf": "132 - Turf",
            "Bed Prep": "154 - Bed Preparation",
            "Bed Preparation": "154 - Bed Preparation",
            "Root Barrier": "26 - Root Barrier",
            "Gravel": "13 - Gravel",
            "Boulders": "29 - Rock & Boulders",
            "River Rock": "29 - Rock & Boulders",
            "Seed": "112 - Seed",
            "Hydroseed": "112 - Seed",
            "Hardscape": "43 - Hardscape",
            "Hardscaping": "43 - Hardscape",
            "Relocate": "136 - Relocate",
            "Relocations": "136 - Relocate",
            "Protect": "142 - Protect",
            "Protection": "142 - Protect",
            "Bed Edging": "154 - Bed Preparation",
            "Replace": "126 - Remove/Replace",
            "Existing Trees to Remain": "1 - Trees & Palms",
            # aliases: estimator vocabulary for a grouping's own word
            "Hedges": "2 - Shrubs",
            "Pavers": "43 - Hardscape",
            "Rip Rap": "29 - Rock & Boulders",
            "Pine Straw": "31 - Mulch (SY)",
            "Sodding": "15 - Sod (SF)",
            "Planting Mix": "4 - Soil",
            "Transplant": "136 - Relocate",
            # Arazoza's own OST group names (from their finished Vela Cove worksheet)
            "01 - Code Trees": "1 - Trees & Palms",
            "02 - Tree Mitigation": "1 - Trees & Palms",
            "03 - Additional Shrubs & Gc": "2 - Shrubs",       # a bare 'Gc' is not expanded — Arazoza files it under Shrubs
            "04 - Code Req.Shrubs": "2 - Shrubs",
            "05 - Bed Areas": "31 - Mulch (SY)",
            "06 - Edging": "154 - Bed Preparation",
            "08 - Soil": "4 - Soil",
            "Remove & Replace": "126 - Remove/Replace",
            "Remove/Replace": "126 - Remove/Replace",
            "Replacements": "27 - Replacements",
        }
        for text, label in cases.items():
            with self.subTest(text=text):
                m = formatter.match_grouping(text, self.groupings)
                self.assertEqual(m.label, label, f"{text!r}: {m.reason}")

    def test_ambiguous_is_flagged(self):
        for text in ("Shrubs & Groundcover", "Trees and Shrubs", "Sod/Turf", "Mulch & Soil", "Gravel/Rock",
                     "Protect Existing Trees", "Tree Protection", "Tree Relocation", "07 - Seed & Sod",
                     # a two-word grouping must not outscore a one-word one
                     "Trees, Palms & Shrubs", "Trees & Palms / Shrubs", "Rock, Boulders & Gravel",
                     "Root Barrier & Mulch", "Bed Preparation & Soil", "Bed Prep & Mulch",
                     # a resemblance must not tip a genuinely two-group header either way
                     "Street Trees & Shrubs", "Ground Cover & Shrubs", "Remove/Replace Trees"):
            with self.subTest(text=text):
                m = formatter.match_grouping(text, self.groupings)
                self.assertIsNone(m.label, f"{text!r} should be ambiguous, got {m.label}")
                self.assertTrue(m.reason.startswith("ambiguous"), m.reason)

    def test_no_match_is_flagged(self):
        for text in ("Irrigation Sleeves", "PLANTING PLAN - SITE", "Landscape", "Annuals", "Grasses", "Lighting",
                     "Site Furnishings", "Vines", "Perennials", "GC",
                     # resemblances that used to relabel confidently — they must be red, not guessed
                     "Streetscape", "Street Furniture", "Phase Three", "Project Totals", "Project Alternates",
                     "Saw Palmetto", "Palmettos", "Bareroot", "Limerock", "Bedrock", "Turfstone", "Hydromulch",
                     "Much", "Oil", "Sold", "Builder", "Shurbs", "Groundcvr"):
            with self.subTest(text=text):
                m = formatter.match_grouping(text, self.groupings)
                self.assertIsNone(m.label, f"{text!r} should not match, got {m.label}")
                self.assertEqual(m.reason, "no match")

    def test_items_beneath_resolve_an_ambiguous_or_unknown_header(self):
        g = self.groupings
        seed_sod = formatter.match_grouping("07 - Seed & Sod", g)
        self.assertIsNone(seed_sod.label)
        placed = formatter.resolve_with_items(seed_sod, ["Sod", "Assumed Sod", "Assumed Sod (Landscape Buffer Area)"], g)
        self.assertEqual((placed.label, placed.reason), ("15 - Sod (SF)", "matched by its items"))
        # items that back BOTH candidates, or neither, leave it ambiguous
        self.assertIsNone(formatter.resolve_with_items(seed_sod, ["Bahia Sod", "Hydroseed"], g).label)
        self.assertIsNone(formatter.resolve_with_items(seed_sod, ["Quercus virginiana / Live Oak"], g).label)
        # an unknown header is rescued only by unanimous items
        misc = formatter.match_grouping("Misc", g)
        self.assertEqual(formatter.resolve_with_items(misc, ['Mulch 3"', "Pine bark mulch"], g).label, "31 - Mulch (SY)")
        self.assertIsNone(formatter.resolve_with_items(misc, ['Mulch 3"', "Sod"], g).label)
        self.assertIsNone(formatter.resolve_with_items(misc, ['Mulch 3"', "Live Oak", "Cabbage Palm"], g).label)
        # a header the text already places is never second-guessed by its items
        trees = formatter.match_grouping("Trees", g)
        self.assertEqual(formatter.resolve_with_items(trees, ["Sod", "Sod"], g).label, "1 - Trees & Palms")

    def test_sod_and_soil_do_not_cross_match(self):
        self.assertEqual(formatter.match_grouping("Sod", self.groupings).label, "15 - Sod (SF)")
        self.assertEqual(formatter.match_grouping("Soil", self.groupings).label, "4 - Soil")
        self.assertEqual(formatter.match_grouping("Trees", self.groupings).label, "1 - Trees & Palms")
        self.assertEqual(formatter.match_grouping("Turf", self.groupings).label, "132 - Turf")


class HelperTests(unittest.TestCase):
    def test_is_black_font(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        cases = [
            (Font(), True),
            (Font(color="FF000000"), True),
            (Font(color="00000000"), True),
            (Font(color=Color(theme=1)), True),
            (Font(color=Color(theme=1, tint=0.0)), True),
            (Font(color=Color(indexed=8)), True),
            (Font(color=Color(indexed=64)), True),
            (Font(color="FFFF0000"), False),          # red
            (Font(color="FF0000FF"), False),          # blue
            (Font(color=Color(theme=0)), False),      # theme 0 is the background colour
            (Font(color=Color(theme=1, tint=0.5)), False),   # grey
            (Font(color=Color(theme=10)), False),     # hyperlink
        ]
        for i, (font, expect) in enumerate(cases, start=1):
            c = ws.cell(i, 1, "x")
            c.font = font
            with self.subTest(i=i, font=font.color):
                self.assertEqual(formatter.is_black_font(c), expect)

    def test_extract_depth(self):
        cases = {
            'Mulch 3" Depth': '3"',
            '3" Depth of Mulch': '3"',
            'Planting Soil Mix 12" Depth': '12"',
            '6" Depth of Top Soil for Bed Areas': '6"',
            "Topsoil 2 in. depth over 500 sf": "2 in.",
            'Mulch at 3-4" depth': '3-4"',
            "Mulch 4' wide ring, 3\" depth": '3"',
            "Soil - 18\" Depth": '18"',
            "Mulch depth": None,
            'Mulch 3"': None,
            "3 individual mulch depth": None,
            # ranges keep both ends, in every spelling
            'Mulch 3"-4" depth': '3"-4"',
            'Soil 6" - 8" depth': '6" - 8"',
            'Mulch 2" to 3" depth': '2" to 3"',
            "Soil 6 in. to 8 in. depth": "6 in. to 8 in.",
            "Soil 1'-6\" depth": "1'-6\"",
            # fractions and mixed numbers
            'Mulch 1/2" depth': '1/2"',
            'Mulch 2 1/2" depth': '2 1/2"',
            'Mulch 1-1/2" depth': '1-1/2"',
            'Mulch 1½" depth': '1½"',
            # hyphenated units
            "Topsoil 3-inch depth": "3-inch",
            "Mulch 24-in depth": "24-in",
        }
        for text, expect in cases.items():
            with self.subTest(text=text):
                self.assertEqual(formatter.extract_depth(text), expect)

    def test_find_packages(self):
        cases = {
            "3 G Cont.": ["Container"],
            "3 G cont": ["Container"],
            "Containers": ["Container"],
            "FG": ["FG"],
            "F.G.": ["FG"],
            "65 gal FG": ["FG"],
            "B&B": ["B&B"],
            "B & B": ["B&B"],
            "FG / B&B": ["FG", "B&B"],
            "Cont, FG, cont": ["Container", "FG"],
            "continuous": [],
            "contractor supplied": [],
            "continuous hedge, cont.": ["Container"],
            "45 G": [],
            "3\" Cal": [],
            "": [],
            # spelled-out / punctuated forms of the same packages
            "Field Grown": ["FG"],
            "field-grown, 6' CT": ["FG"],
            "B.&B.": ["B&B"],
            "B&Bs": ["B&B"],
            "balled and burlapped": ["B&B"],
            "container grown": ["Container"],
            "finish grade": [],
        }
        for text, expect in cases.items():
            with self.subTest(text=text):
                self.assertEqual(formatter.find_packages(text), expect)

    def test_add_package(self):
        self.assertEqual(formatter._add_package(None, "Depth"), "Depth")
        self.assertEqual(formatter._add_package("-", "Depth"), "Depth")
        self.assertEqual(formatter._add_package("Container", "FG"), "Container/FG")
        self.assertEqual(formatter._add_package("Container", "container"), "Container")

    def test_output_name(self):
        self.assertEqual(formatter.output_name_for("Arazoza - X - Worksheet 2026-09-04.xlsm"),
                         "Arazoza - X - Worksheet 2026-09-04 - formatted.xlsm")
        self.assertEqual(formatter.output_name_for("/tmp/file_1-Arazoza.XLSX"), "file_1-Arazoza - formatted.xlsx")


class SyntheticEndToEndTests(unittest.TestCase):
    """Every row of make_fixture.SYNTHETIC_ROWS, checked after the full seven-step run."""

    @classmethod
    def setUpClass(cls):
        cls.tmp = tempfile.TemporaryDirectory()
        raw = make_fixture.build_synthetic_raw(Path(cls.tmp.name) / "raw.xlsx")
        cls.out = Path(cls.tmp.name) / "out.xlsx"
        cls.report = formatter.format_file(raw, cls.out)
        cls.ws = openpyxl.load_workbook(cls.out)["Project Totals"]

    @classmethod
    def tearDownClass(cls):
        cls.tmp.cleanup()

    def cell(self, ref):
        return self.ws[ref]

    def val(self, ref):
        return self.ws[ref].value

    def test_picked_the_visible_data_sheet_not_the_hidden_decoy(self):
        self.assertEqual(self.report.sheet, "Project Totals")
        self.assertEqual(self.report.header_row, 15)
        self.assertEqual(self.report.columns, {"description": "B", "size": "C", "package": "D", "notes": "E", "qty": "F", "uom": "G"})

    # step 1 ------------------------------------------------------------------------
    def test_step1_black_headers_relabelled(self):
        self.assertEqual(self.val("B18"), "1 - Trees & Palms")
        self.assertEqual(self.val("B22"), "1 - Trees & Palms")
        self.assertEqual(self.val("B28"), "31 - Mulch (SY)")
        self.assertEqual(self.val("B32"), "4 - Soil")
        self.assertEqual(self.val("B36"), "154 - Bed Preparation")
        self.assertEqual(self.val("B39"), "15 - Sod (SF)")
        self.assertEqual(self.val("B49"), "1 - Trees & Palms")
        for ref in ("B18", "B22", "B28", "B32", "B36", "B39", "B49"):
            self.assertFalse(_red(self.cell(ref)), ref)

    def test_step1_already_final_header_untouched_and_counted(self):
        self.assertEqual(self.val("B43"), "1 - Trees & Palms")
        self.assertIn((43, "1 - Trees & Palms"), self.report.groups_unchanged)

    def test_step1_coloured_section_marker_skipped(self):
        self.assertEqual(self.val("B16"), "Landscape")
        self.assertFalse(_red(self.cell("B16")))
        self.assertIn((16, "Landscape"), self.report.groups_skipped_colored)

    def test_step1_unmatched_and_ambiguous_filled_red_and_text_kept(self):
        for ref, text in (("B25", "Shrubs & Groundcover"), ("B41", "Irrigation Sleeves"), ("B46", "PLANTING PLAN - SITE")):
            with self.subTest(ref=ref):
                self.assertEqual(self.val(ref), text)
                self.assertTrue(_red(self.cell(ref)))
        reasons = {row: reason for row, _, reason in self.report.groups_flagged}
        self.assertTrue(reasons[25].startswith("ambiguous"), reasons[25])
        self.assertIn("2 - Shrubs", reasons[25])
        self.assertIn("9 - Groundcover", reasons[25])
        self.assertEqual(reasons[41], "no match")
        self.assertEqual(reasons[46], "no match")

    # step 2 ------------------------------------------------------------------------
    def test_summary_says_when_it_held_notes_back(self):
        many = formatter.Report(sheet="S", warnings=[f"w{i}" for i in range(9)])
        text = many.summary("a.xlsx", "b.xlsx")
        self.assertIn("…and 3 more notes", text)
        self.assertNotIn("…and", self.report.summary("a.xlsx", "b.xlsx").split("Note:")[0])

    def test_step2_items_moved_from_size_to_description_and_size_cleared_or_refilled(self):
        self.assertEqual(self.val("B30"), "Assumed @ 4' Dia of Mulch Ring")
        self.assertIsNone(self.val("C30"))                    # nothing in Notes -> Size ends empty
        self.assertEqual(self.val("B37"), "Bed Edging")
        self.assertIsNone(self.val("C37"))
        self.assertEqual(self.report.items_moved, 26)     # every C value except the conflict row

    def test_step2_conflict_row_flagged_and_left_entirely_alone(self):
        self.assertEqual(self.val("B45"), "Root Barrier")      # not relabelled, not merged
        self.assertEqual(self.val("C45"), "Conflict item, cont")
        self.assertIsNone(self.val("D45"))                     # not even the 'cont' in its Size
        self.assertEqual(self.val("G45"), "EA")                # not even its UOM
        self.assertTrue(_red(self.cell("B45")))
        self.assertEqual(self.report.conflicts, [(45, "Root Barrier", "Conflict item, cont")])
        self.assertNotIn(45, [r for r, _, _ in self.report.groups_matched])
        self.assertNotIn(45, [r for r, _ in self.report.packages_set])

    # steps 3 + 4 -------------------------------------------------------------------
    def test_step3_4_notes_become_size_and_are_appended_to_description(self):
        self.assertEqual(self.val("C19"), "45 G, 12'-14' Ht. x 6'-7' Sp., 3\" Cal")
        self.assertIsNone(self.val("E19"))
        self.assertEqual(self.val("B19"), "Quercus virginiana / Live Oak - 45 G, 12'-14' Ht. x 6'-7' Sp., 3\" Cal")
        self.assertEqual(self.val("B26"), "Clusia guttifera / Small Leaf Clusia - 3 G, 36\" Ht. x 24\" Sp.")
        self.assertEqual(self.val("B42"), "4\" PVC Sleeve - Sch 40")
        self.assertEqual(self.val("C42"), "Sch 40")

    def test_step4_size_already_in_description_is_not_appended_twice(self):
        self.assertEqual(self.val("B27"), "Ficus microcarpa Green Island - 3 G")
        self.assertEqual(self.val("C27"), "3 G")
        self.assertEqual([r for r, _, _ in self.report.merge_skipped], [27])

    def test_step4_only_an_exact_suffix_counts_as_already_there(self):
        # 'Cont' is a substring of Contorta and '3' of many names — the size must still be appended
        with tempfile.TemporaryDirectory() as tmp:
            wb = openpyxl.Workbook(); ws = wb.active
            for i, h in enumerate(make_fixture.HEADERS):
                ws.cell(1, 2 + i, h)
            ws["C2"], ws["E2"] = "Ilex 3-5 pips", 3
            ws["C3"], ws["E3"] = "Corylus avellana 'Contorta'", "Cont"
            src = Path(tmp) / "s.xlsx"; wb.save(src)
            formatter.format_file(src, Path(tmp) / "o.xlsx")
            out = openpyxl.load_workbook(Path(tmp) / "o.xlsx").active
            self.assertEqual(out["B2"].value, "Ilex 3-5 pips - 3")
            self.assertEqual(out["B3"].value, "Corylus avellana 'Contorta' - Cont")
            self.assertEqual(out["D3"].value, "Container")

    def test_step3_numeric_notes_value_moves_as_is_and_merges_as_text(self):
        self.assertEqual(self.val("C48"), 3)
        self.assertEqual(self.val("B48"), "Numeric Note Item - 3")

    def test_step3_date_notes_value_is_merged_as_a_date_and_flagged_in_the_report(self):
        self.assertEqual(self.val("B57"), "Date Note Item - 2023-03-04")
        self.assertTrue(any("row 57" in w and "date" in w for w in self.report.warnings), self.report.warnings)

    def test_header_rows_keep_their_notes_and_are_never_merged(self):
        # a matched header with a note, and a coloured marker with a note: label/marker untouched,
        # note left in Notes, nothing in Size or Package
        self.assertEqual(self.val("B51"), "31 - Mulch (SY)")
        self.assertEqual(self.val("E51"), '3" depth cont')
        self.assertIsNone(self.val("C51"))
        self.assertIsNone(self.val("D51"))
        self.assertEqual(self.val("B55"), "IRRIGATION")
        self.assertEqual(self.val("E55"), "per IR-1")
        self.assertIsNone(self.val("C55"))
        self.assertFalse(_red(self.cell("B55")))

    # step 5 ------------------------------------------------------------------------
    def test_step5_depth_from_notes_after_merge(self):
        # raw: Size 'Mulch', Notes '3" Depth' -> merged 'Mulch - 3" Depth' -> Size 3", Package Depth
        self.assertEqual(self.val("B29"), 'Mulch - 3" Depth')
        self.assertEqual(self.val("C29"), '3"')
        self.assertEqual(self.val("D29"), "Depth")

    def test_step5_depth_in_item_text_with_no_notes(self):
        self.assertEqual(self.val("B33"), 'Planting Soil Mix 12" Depth')
        self.assertEqual(self.val("C33"), '12"')
        self.assertEqual(self.val("D33"), "Depth")
        self.assertEqual(self.val("B35"), 'Topsoil for Grass Areas - 2" Depth')
        self.assertEqual(self.val("C35"), '2"')
        self.assertEqual(self.val("D35"), "Depth")

    def test_step5_soil_recognised_by_group_header_not_only_by_name(self):
        self.assertEqual(self.val("C34"), '6"')
        self.assertEqual(self.val("D34"), "Depth")

    def test_step5_generic_size_without_depth_word_left_alone(self):
        self.assertEqual(self.val("B31"), 'Pine Bark Mulch - 3"')
        self.assertEqual(self.val("C31"), '3"')
        self.assertIsNone(self.val("D31"))

    def test_step5_depth_word_on_non_soil_mulch_item_ignored(self):
        self.assertEqual(self.val("B38"), 'Weed Barrier 3" depth')
        self.assertIsNone(self.val("C38"))
        self.assertIsNone(self.val("D38"))

    def test_step5_depth_word_without_measurement_reported_not_changed(self):
        self.assertEqual(self.val("B47"), "Depth of mulch unknown")
        self.assertIsNone(self.val("C47"))
        self.assertIsNone(self.val("D47"))
        self.assertEqual([r for r, _ in self.report.depth_unparsed], [47])

    def test_step5_runs_before_step7_so_depth_wins_over_a_package_token(self):
        self.assertEqual(self.val("B52"), 'Pine Straw Mulch - bulk cont., 3" depth')
        self.assertEqual(self.val("C52"), '3"')
        self.assertEqual(self.val("D52"), "Depth")            # not 'Container/Depth'

    def test_step5_soil_mulch_context_stops_at_a_flagged_or_coloured_header(self):
        # 'Crushed Granite' sits under the flagged 'Decorative Stone', which follows the Mulch
        # group — it is NOT a mulch item, so its depth stays in the text and its package is read
        self.assertTrue(_red(self.cell("B53")))
        self.assertEqual(self.val("B54"), 'Crushed Granite - 4" depth, cont')
        self.assertEqual(self.val("C54"), '4" depth, cont')
        self.assertEqual(self.val("D54"), "Container")
        # and nothing under the red 'IRRIGATION' marker is soil or mulch either
        self.assertEqual(self.val("B56"), '2" PVC Mainline - Sch 40, 18" depth cover')
        self.assertEqual(self.val("C56"), 'Sch 40, 18" depth cover')
        self.assertIsNone(self.val("D56"))

    def test_step5_fraction_depth(self):
        self.assertEqual(self.val("C58"), '2 1/2"')
        self.assertEqual(self.val("D58"), "Depth")

    def test_a_description_row_carrying_a_qty_is_an_item_not_a_header(self):
        # Its name must survive step 1 (a header match would have overwritten row 59 with
        # '31 - Mulch (SY)' and row 60 with '1 - Trees & Palms')...
        self.assertEqual(self.val("B59"), 'Pine Straw Mulch 3" depth')
        self.assertEqual(self.val("B60"), "Existing Tree to Remain")
        for ref in ("B59", "B60"):
            self.assertFalse(_red(self.cell(ref)), ref)
        self.assertNotIn(59, [r for r, _, _ in self.report.groups_matched])
        self.assertNotIn(60, [r for r, _, _ in self.report.groups_matched])
        # ...and the item steps still run on it: 'pine straw' is mulch, so the depth applies
        self.assertEqual(self.val("C59"), '3"')
        self.assertEqual(self.val("D59"), "Depth")
        self.assertEqual(self.val("G59"), "Square Feet")
        self.assertIsNone(self.val("C60"))
        self.assertEqual(self.val("G60"), "Unit")

    def test_depth_without_a_package_column_is_reported_not_silently_dropped(self):
        with tempfile.TemporaryDirectory() as tmp:
            wb = openpyxl.Workbook(); ws = wb.active
            for i, h in enumerate(["Description", "Size", "Notes", "Qty 1", "UOM"]):   # no Package
                ws.cell(1, 2 + i, h)
            ws["C2"], ws["D2"], ws["E2"], ws["F2"] = "Mulch", '3" depth', 100, "SF"
            src = Path(tmp) / "nopkg.xlsx"; wb.save(src)
            report = formatter.format_file(src, Path(tmp) / "out.xlsx")
            out = openpyxl.load_workbook(Path(tmp) / "out.xlsx").active
            self.assertEqual(out["C2"].value, '3"')
            self.assertEqual([r for r, _, _, _ in report.depth_rows], [2])
            self.assertTrue(any("no 'Package' column" in w for w in report.warnings), report.warnings)
            self.assertIn("no 'Package' column", report.summary("nopkg.xlsx", "out.xlsx"))

    # step 6 ------------------------------------------------------------------------
    def test_step6_uom_conversions_primary_column_only(self):
        expect = {"G19": "Unit", "G20": "Unit", "G21": "Unit", "G23": "Unit", "G24": "Unit", "G26": "Unit",
                  "G29": "Square Feet", "G31": "Square Feet", "G33": "Square Feet", "G37": "Linear Feet",
                  "G38": "Linear Feet", "G40": "Square Feet", "G42": "Linear Feet", "G44": "CY", "G47": "Unit",
                  "G52": "Square Feet", "G54": "Square Feet", "G56": "Linear Feet", "G58": "Square Feet"}
        for ref, v in expect.items():
            with self.subTest(ref=ref):
                self.assertEqual(self.val(ref), v)
        # the secondary UOM column keeps its raw values
        self.assertEqual(self.val("I19"), "SF")
        self.assertEqual(self.val("I29"), "CY")
        self.assertEqual(self.val("I40"), "CY")

    # step 7 ------------------------------------------------------------------------
    def test_step7_packages_copied_from_size(self):
        self.assertEqual(self.val("D20"), "FG")
        self.assertEqual(self.val("D21"), "Container")
        self.assertEqual(self.val("D23"), "B&B")
        self.assertEqual(self.val("D50"), "Container")
        self.assertIsNone(self.val("D19"))                    # '45 G' is not a package token
        self.assertIsNone(self.val("D26"))

    def test_step7_two_packages_joined_with_slash_and_merged_with_existing(self):
        self.assertEqual(self.val("D24"), "Container/FG/B&B")

    def test_step7_depth_rows_keep_depth_package(self):
        self.assertEqual(self.val("D29"), "Depth")

    # untouched ---------------------------------------------------------------------
    def test_other_columns_untouched(self):
        self.assertEqual(self.val("F19"), 68)
        self.assertEqual(self.val("H19"), 1900)
        self.assertEqual(self.val("J50"), 5)
        self.assertEqual(self.val("B15"), "Description")
        self.assertEqual(self.val("B1"), "Arazoza Brothers")

    def test_summary_mentions_what_matters(self):
        s = self.report.summary("raw.xlsx", "raw - formatted.xlsx")
        self.assertIn("RED", s)
        self.assertIn("row 25", s)
        self.assertIn("Irrigation Sleeves", s)
        self.assertIn("row 45", s)
        self.assertIn("'Trees' → '1 - Trees & Palms'", s)
        self.assertIn("UOM converted", s)
        self.assertIn("Packages filled", s)
        self.assertLess(len(s), 3000)


class SafetyTests(unittest.TestCase):
    def test_refuses_already_formatted_sheet_unless_forced(self):
        with tempfile.TemporaryDirectory() as tmp:
            src = make_fixture.build_already_formatted(Path(tmp) / "done.xlsx")
            with self.assertRaises(formatter.FormatterError) as ctx:
                formatter.format_file(src, Path(tmp) / "out.xlsx")
            self.assertIn("nothing in 'Size' to move", str(ctx.exception))
            self.assertFalse((Path(tmp) / "out.xlsx").exists())
            report = formatter.format_file(src, Path(tmp) / "out2.xlsx", force=True)
            self.assertTrue((Path(tmp) / "out2.xlsx").exists())
            # ...and this is why the guard exists: on a finished sheet every item has BOTH
            # Description and Size filled, so each one is a conflict row — red-filled, untouched.
            self.assertEqual([r for r, _, _ in report.conflicts], [17, 18])
            self.assertEqual(report.groups_flagged, [])
            self.assertEqual(report.groups_matched, [])
            self.assertEqual(report.items_moved, 0)

    def test_rejects_unsupported_extension(self):
        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "old.xls"
            p.write_bytes(b"not really")
            with self.assertRaises(formatter.FormatterError):
                formatter.format_file(p, Path(tmp) / "out.xls")

    def test_missing_header_row_is_a_clear_error(self):
        with tempfile.TemporaryDirectory() as tmp:
            wb = openpyxl.Workbook()
            wb.active["A1"] = "just some text"
            p = Path(tmp) / "nohdr.xlsx"
            wb.save(p)
            with self.assertRaises(formatter.FormatterError) as ctx:
                formatter.format_file(p, Path(tmp) / "out.xlsx")
            self.assertIn("'Description' and a 'Size' header", str(ctx.exception))

    def test_unknown_sheet_name_is_a_clear_error(self):
        with tempfile.TemporaryDirectory() as tmp:
            src = make_fixture.build_synthetic_raw(Path(tmp) / "raw.xlsx")
            with self.assertRaises(formatter.FormatterError) as ctx:
                formatter.format_file(src, Path(tmp) / "out.xlsx", sheet_name="Nope")
            self.assertIn("no sheet called 'Nope'", str(ctx.exception))

    def test_original_file_is_not_modified(self):
        with tempfile.TemporaryDirectory() as tmp:
            src = make_fixture.build_synthetic_raw(Path(tmp) / "raw.xlsx")
            before = src.read_bytes()
            formatter.format_file(src, Path(tmp) / "out.xlsx")
            self.assertEqual(src.read_bytes(), before)


class RunPyContractTests(unittest.TestCase):
    """run.py end to end, the way tool_runner.LocalSubprocessBackend calls it."""

    def test_ok_result_and_artifact(self):
        with tempfile.TemporaryDirectory() as tmp:
            staged = Path(tmp) / "file_1-Arazoza - Fixture - Worksheet 2026-09-04.xlsx"
            make_fixture.build_synthetic_raw(staged)
            work = Path(tmp) / "work"
            work.mkdir()
            contract = {"input": {"input_file": "file_1"}, "input_path": str(staged), "work_dir": str(work), "backend": "local"}
            proc = subprocess.run([sys.executable, str(TOOL_DIR / "run.py")], cwd=str(TOOL_DIR),
                                  input=json.dumps(contract), capture_output=True, text=True, timeout=120)
            self.assertEqual(proc.returncode, 0, proc.stderr)
            result = json.loads((work / "result.json").read_text())
            self.assertEqual(result["status"], "ok", result)
            self.assertEqual(len(result["artifacts"]), 1)
            art = result["artifacts"][0]
            self.assertEqual(art["filename"], "Arazoza - Fixture - Worksheet 2026-09-04 - formatted.xlsx")
            self.assertEqual(art["kind"], "xlsx")
            self.assertTrue(Path(art["ref"]).exists())
            self.assertIn("RED", result["summary"])
            self.assertEqual(json.loads(proc.stdout.strip().splitlines()[-1])["status"], "ok")

    def test_error_result_for_formatted_sheet(self):
        with tempfile.TemporaryDirectory() as tmp:
            staged = Path(tmp) / "file_1-done.xlsx"
            make_fixture.build_already_formatted(staged)
            work = Path(tmp) / "work"
            work.mkdir()
            contract = {"input": {"input_file": "file_1"}, "input_path": str(staged), "work_dir": str(work), "backend": "local"}
            proc = subprocess.run([sys.executable, str(TOOL_DIR / "run.py")], cwd=str(TOOL_DIR),
                                  input=json.dumps(contract), capture_output=True, text=True, timeout=120)
            result = json.loads((work / "result.json").read_text())
            self.assertEqual(result["status"], "error")
            self.assertIn("already", result["error"])
            self.assertEqual(result["artifacts"], [])


@unittest.skipUnless(VELA_RAW and VELA_FINISHED, "Vela Cove OST export + finished worksheet not on this machine")
class VelaCoveGoldenTests(unittest.TestCase):
    """The real thing: an Arazoza worksheet straight out of OST, and the same worksheet after a
    human formatted it. The tool must land on the human's result everywhere except three
    documented classes of difference that the spec does not cover:
      D  the human types '-' for trees/palms and 'Container' for gallon sizes (the spec's package
         rule only lifts FG / cont / container / B&B out of Size);
      G  the human changed OST's 'LF' to 'Unit' on shrub rows (the spec says lf -> Linear Feet);
      B  two descriptions where the human typed extra words by hand."""

    @classmethod
    def setUpClass(cls):
        cls.tmp = tempfile.TemporaryDirectory()
        cls.out = Path(cls.tmp.name) / formatter.output_name_for(VELA_RAW)
        cls.report = formatter.format_file(VELA_RAW, cls.out)
        cls.tool = openpyxl.load_workbook(cls.out, keep_vba=True)["Project Totals"]
        cls.human = openpyxl.load_workbook(VELA_FINISHED, keep_vba=True)["Project Totals"]
        # the human inserted a blank row after 'Landscape'; align the two sheets by data rows
        data = lambda ws: [r for r in range(16, ws.max_row + 1) if formatter._txt(ws.cell(r, 2).value)]
        cls.pairs = list(zip(data(cls.tool), data(cls.human)))

    @classmethod
    def tearDownClass(cls):
        cls.tmp.cleanup()

    def test_same_number_of_rows(self):
        self.assertEqual(len(self.pairs), 57)

    def test_every_group_header_lands_on_the_humans_label(self):
        headers = [(a, self.tool.cell(a, 2).value, self.human.cell(b, 2).value)
                   for a, b in self.pairs if self.tool.cell(a, 2).font.b]
        self.assertEqual(len(headers), 9)                     # 'Landscape' + 8 groups
        for row, got, want in headers:
            with self.subTest(row=row):
                self.assertEqual(got, want)
        self.assertEqual(sorted(t for _, t, _ in self.report.groups_matched),
                         sorted(["01 - Code Trees", "02 - Tree Mitigation", "03 - Additional Shrubs & Gc", "04 - Code Req.Shrubs",
                                 "05 - Bed Areas", "06 - Edging", "07 - Seed & Sod", "08 - Soil"]))
        self.assertEqual([t for _, t, _ in self.report.groups_by_items], ["07 - Seed & Sod"])

    def test_nothing_is_flagged_red(self):
        self.assertEqual(self.report.groups_flagged, [])
        self.assertEqual(self.report.conflicts, [])
        for a, _ in self.pairs:
            self.assertFalse(_red(self.tool.cell(a, 2)), f"row {a} is red")

    def test_descriptions_sizes_depths_and_uoms_match_the_human(self):
        import re
        squash = lambda t: re.sub(r"\s+", " ", t)
        human_extras = {"3\" Depth of Mulch": "3\" Depth of Mulch - Bulk Organic Mulch",
                        "4\"-6\" Size, 10\" Depth of Riprap Rock Bed": "4\"-6\" Size, 10\" Depth of Riprap Rock Bed/W weed Fabric"}
        diffs = []
        for a, b in self.pairs:
            tb, hb = formatter._txt(self.tool.cell(a, 2).value), formatter._txt(self.human.cell(b, 2).value)
            if squash(tb) != squash(hb) and human_extras.get(tb) != hb:
                diffs.append(("B", a, tb, hb))
            tc, hc = self.tool.cell(a, 3).value, self.human.cell(b, 3).value
            if (tc or "") != (hc or ""):                       # exact: the human kept OST's leading spaces, so do we
                diffs.append(("C", a, tc, hc))
            td, hd = formatter._txt(self.tool.cell(a, 4).value), formatter._txt(self.human.cell(b, 4).value)
            if hd == "Depth" and td != "Depth":
                diffs.append(("D", a, td, hd))
            if hd not in ("Depth", "-", "Container") and td != hd:
                diffs.append(("D", a, td, hd))
            tg, hg = formatter._txt(self.tool.cell(a, 7).value), formatter._txt(self.human.cell(b, 7).value)
            if tg != hg and not (tg == "Linear Feet" and hg == "Unit"):
                diffs.append(("G", a, tg, hg))
        self.assertEqual(diffs, [], "\n".join(map(str, diffs)))

    def test_the_documented_differences_are_exactly_those(self):
        # the human's LF -> Unit override is confined to the two shrub groups
        lf_rows = [a for a, b in self.pairs
                   if formatter._txt(self.tool.cell(a, 7).value) == "Linear Feet" and formatter._txt(self.human.cell(b, 7).value) == "Unit"]
        self.assertEqual(len(lf_rows), 19)
        self.assertTrue(all(40 <= r <= 59 for r in lf_rows), lf_rows)
        # and the tool never writes a package the spec doesn't define
        self.assertEqual({formatter._txt(self.tool.cell(a, 4).value) for a, _ in self.pairs}, {"", "Depth"})

    def test_workbook_survives_including_the_extra_sheet(self):
        names = zipfile.ZipFile(self.out).namelist()
        self.assertIn("xl/vbaProject.bin", names)
        wb = openpyxl.load_workbook(self.out, keep_vba=True)
        self.assertEqual(wb.sheetnames, openpyxl.load_workbook(VELA_RAW, keep_vba=True).sheetnames)
        self.assertIn("Sheet1", wb.sheetnames)
        self.assertEqual(dict(self.tool.tables.items()), {"LH_materials_list_tbl": "B15:T140"})


class LambdaHandlerTests(unittest.TestCase):
    """lambda/handler.py against a fake S3: same contract as run.py, files by key."""

    class FakeS3:
        def __init__(self, store):
            self.store = store
        def download_file(self, bucket, key, path):
            Path(path).write_bytes(self.store[key])
        def put_object(self, Bucket, Key, Body):
            self.store[Key] = Body

    def _handler(self):
        sys.path.insert(0, str(TOOL_DIR / "lambda"))
        import importlib
        import handler
        return importlib.reload(handler)

    def test_ok_round_trip_by_s3_key(self):
        handler = self._handler()
        with tempfile.TemporaryDirectory() as tmp:
            raw = make_fixture.build_synthetic_raw(Path(tmp) / "raw.xlsx")
            key = "runs/abc/input/file_1-Arazoza - Fixture - Worksheet 2026-09-04.xlsx"
            store = {key: raw.read_bytes()}
            handler.s3 = self.FakeS3(store)
            res = handler.handler({"input": {"input_file": "file_1"}, "input_path": key,
                                   "work_dir": "runs/abc/work-1/output", "bucket": "b", "backend": "lambda"}, None)
            self.assertEqual(res["status"], "ok", res)
            art = res["artifacts"][0]
            self.assertEqual(art["ref"], "runs/abc/work-1/output/Arazoza - Fixture - Worksheet 2026-09-04 - formatted.xlsx")
            self.assertEqual(art["filename"], "Arazoza - Fixture - Worksheet 2026-09-04 - formatted.xlsx")
            self.assertIn(art["ref"], store)
            out = openpyxl.load_workbook(__import__("io").BytesIO(store[art["ref"]]))["Project Totals"]
            self.assertEqual(out["B18"].value, "1 - Trees & Palms")
            self.assertIn("RED", res["summary"])

    def test_error_path_is_a_result_not_an_exception(self):
        handler = self._handler()
        with tempfile.TemporaryDirectory() as tmp:
            done = make_fixture.build_already_formatted(Path(tmp) / "done.xlsx")
            key = "runs/abc/input/file_1-done.xlsx"
            handler.s3 = self.FakeS3({key: done.read_bytes()})
            res = handler.handler({"input": {}, "input_path": key, "work_dir": "runs/abc/w/output", "bucket": "b"}, None)
            self.assertEqual(res["status"], "error")
            self.assertIn("already", res["error"])
            res = handler.handler({"input": {}, "input_path": None, "work_dir": "x", "bucket": "b"}, None)
            self.assertEqual(res["status"], "error")


@unittest.skipUnless(FINISHED_SAMPLE.exists(), f"real sample not on this machine: {FINISHED_SAMPLE}")
class RealSampleTests(unittest.TestCase):
    """Reverse a finished, human-made Arazoza .xlsm into the raw layout, run the tool, and
    check it rebuilds the finished sheet — and that the macro workbook comes back intact."""

    @classmethod
    def setUpClass(cls):
        cls.tmp = tempfile.TemporaryDirectory()
        cls.raw = Path(cls.tmp.name) / "Arazoza - Esplanade RAW.xlsm"
        cls.expected = make_fixture.deformat_finished(FINISHED_SAMPLE, cls.raw)["expected"]
        cls.out = Path(cls.tmp.name) / formatter.output_name_for(cls.raw)
        cls.report = formatter.format_file(cls.raw, cls.out)
        cls.wb = openpyxl.load_workbook(cls.out, keep_vba=True)
        cls.ws = cls.wb["Project Totals"]

    @classmethod
    def tearDownClass(cls):
        cls.tmp.cleanup()

    def test_descriptions_sizes_and_uoms_match_the_human_made_sheet(self):
        mismatches = []
        for row, exp in self.expected.items():
            got_b = formatter._txt(self.ws.cell(row, 2).value)
            got_c = formatter._txt(self.ws.cell(row, 3).value)
            got_g = formatter._txt(self.ws.cell(row, 7).value)
            if (got_b, got_c, got_g) != (exp["B"], exp["C"], exp["G"]):
                mismatches.append((row, (got_b, got_c, got_g), (exp["B"], exp["C"], exp["G"])))
        self.assertEqual(mismatches, [], "\n".join(map(str, mismatches)))

    def test_depth_rows_get_size_and_package_like_the_human_did(self):
        depth_rows = {row: exp for row, exp in self.expected.items() if exp["D"] == "Depth"}
        self.assertGreaterEqual(len(depth_rows), 2)
        for row, exp in depth_rows.items():
            with self.subTest(row=row):
                self.assertEqual(formatter._txt(self.ws.cell(row, 3).value), exp["C"])
                self.assertEqual(formatter._txt(self.ws.cell(row, 4).value), "Depth")

    def test_package_only_from_declared_tokens(self):
        # The human-made sheet says 'Container' for every gallon size; the spec's package rule
        # only lifts FG / cont / container / B&B out of Size, and none of these sizes carry one,
        # so Package stays empty — a documented difference, not a bug.
        for row, exp in self.expected.items():
            if exp["D"] in ("Container", "-"):
                self.assertEqual(formatter._txt(self.ws.cell(row, 4).value), "", f"row {row}")

    def test_nothing_flagged_red_on_a_clean_sheet(self):
        self.assertEqual(self.report.groups_flagged, [])
        self.assertEqual(self.report.conflicts, [])
        self.assertEqual(self.report.groups_skipped_colored, [(16, "Landscape")])

    def test_macro_workbook_survives(self):
        names = zipfile.ZipFile(self.out).namelist()
        self.assertIn("xl/vbaProject.bin", names)
        self.assertTrue(any(n.startswith("customUI/") for n in names))
        src = openpyxl.load_workbook(FINISHED_SAMPLE, keep_vba=True)
        self.assertEqual(self.wb.sheetnames, src.sheetnames)
        self.assertEqual([s.sheet_state for s in self.wb.worksheets], [s.sheet_state for s in src.worksheets])
        self.assertEqual(dict(self.ws.tables.items()), {"LH_materials_list_tbl": "B15:T140"})
        # the import formulas in K..T are carried through verbatim
        for row in (16, 40, 140):
            for col in range(11, 21):
                self.assertEqual(self.ws.cell(row, col).value, src["Project Totals"].cell(row, col).value)
        self.assertEqual(self.ws["J19"].value, src["Project Totals"]["J19"].value)   # Column1 untouched

    def test_refuses_the_finished_sample_itself(self):
        with self.assertRaises(formatter.FormatterError):
            formatter.format_file(FINISHED_SAMPLE, Path(self.tmp.name) / "should-not-exist.xlsm")

    def test_rows_pasted_below_the_excel_table_are_formatted_and_reported(self):
        raw2 = Path(self.tmp.name) / "below.xlsm"
        make_fixture.deformat_finished(FINISHED_SAMPLE, raw2)
        wb = openpyxl.load_workbook(raw2, keep_vba=True); ws = wb["Project Totals"]
        ws["B141"] = "Groundcover"
        ws["C142"], ws["E142"], ws["F142"], ws["G142"] = "Liriope muscari / Lilyturf", "1 G", 500, "EA"
        ws["C143"], ws["E143"], ws["F143"], ws["G143"] = "Beyond-table item", "3 G cont", 10, "ea"
        wb.save(raw2)
        out2 = Path(self.tmp.name) / "below-out.xlsm"
        report = formatter.format_file(raw2, out2)
        ws = openpyxl.load_workbook(out2, keep_vba=True)["Project Totals"]
        self.assertEqual(ws["B141"].value, "9 - Groundcover")
        self.assertEqual(ws["B142"].value, "Liriope muscari / Lilyturf - 1 G")
        self.assertEqual(ws["B143"].value, "Beyond-table item - 3 G cont")
        self.assertEqual(ws["D143"].value, "Container")
        self.assertEqual(ws["G143"].value, "Unit")
        self.assertTrue(any("141" in w and "table" in w for w in report.warnings), report.warnings)


@unittest.skipUnless(SAMPLE_DIR.is_dir(), "real samples not on this machine")
class AllLandscapeSamplesTests(unittest.TestCase):
    """The reverse-and-format round trip must hold for every landscape Arazoza sample, not just
    the hand-picked one. Irrigation-only sheets are excluded: their headers aren't in the
    (landscape-only) groupings list and go red by design."""

    LANDSCAPE = (
        "Arazoza - Esplanade at Tradition Ph 4 - Worksheet 2023-10-09.xlsm",
        "Arazoza - Liberty Square Ph IV Rev 2-A - Worksheet 2023-04-27.xlsm",
        "Arazoza - Royal Pointe - Worksheet 2023-10-06.xlsm",
        "Arazoza - 61st Street Townhomes - Worksheet 2022-05-17.xlsm",
    )

    def test_every_landscape_sample_round_trips(self):
        present = [n for n in self.LANDSCAPE if (SAMPLE_DIR / n).exists()]
        self.assertGreaterEqual(len(present), 1, f"no landscape samples found in {SAMPLE_DIR}")
        for name in present:
            src = SAMPLE_DIR / name
            with self.subTest(sample=name), tempfile.TemporaryDirectory() as tmp:
                raw = Path(tmp) / "raw.xlsm"
                expected = make_fixture.deformat_finished(src, raw)["expected"]
                out = Path(tmp) / "out.xlsm"
                report = formatter.format_file(raw, out)
                ws = openpyxl.load_workbook(out, keep_vba=True)["Project Totals"]
                import re
                squash = lambda t: re.sub(r"\s+", " ", t)      # a human's stray double space is not a defect
                mismatches = []
                for row, exp in expected.items():
                    got = (squash(formatter._txt(ws.cell(row, 2).value)), formatter._txt(ws.cell(row, 3).value),
                           formatter._txt(ws.cell(row, 7).value))
                    if got != (squash(exp["B"]), exp["C"], exp["G"]):
                        mismatches.append((row, got, (exp["B"], exp["C"], exp["G"])))
                self.assertEqual(mismatches, [], "\n".join(map(str, mismatches)))
                self.assertGreater(len(expected), 20, name)
                # only the plan-section sub-headers a human wrote in black may be red
                for _, text, _ in report.groups_flagged:
                    self.assertIn("PLAN", text.upper(), f"{name}: unexpected red header {text!r}")


if __name__ == "__main__":
    unittest.main()
