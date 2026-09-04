"""Arazoza worksheet formatter — the logic behind the `arazoza-formatter` bot tool.

Arazoza Brothers' takeoff worksheets arrive in a layout that is *almost* the Takeoff Monkey
worksheet, but with the data one column to the right of where it belongs: the item name sits
in `Size`, the sizing sits in `Notes`, and the group headers in `Description` carry Arazoza's
own names rather than our numbered takeoff groupings. This module walks the sheet through a
fixed sequence of moves that puts everything where the worksheet (and the downstream import)
expects it, and paints red anything a human still has to look at.

THE ORDER OF THE STEPS IS THE SPEC (TODO #6). Each step reads what the previous one wrote:
the group matching has to run while `Description` still holds only the headers, the ' - '
merge only after `Notes` has become `Size`, the depth rule only after the merge, and so on.
Reordering them yields plausible-looking but wrong data, so `format_worksheet` is the single
place that sequences them and nothing else calls the steps directly.

  1. Match every BLACK-text `Description` cell to a takeoff grouping (takeoff-groupings.txt);
     no match or an ambiguous match -> the cell is filled red.
  2. Move every `Size` value into the (empty) `Description` cell on its row.
  3. Move every `Notes` value into the (now empty) `Size` cell on its row.
  4. Description = Description + " - " + Size (where both exist).
  5. Soil/mulch items whose text says "depth": Size = the depth measurement, Package += Depth.
  6. UOM: count/ea/each -> Unit, sf -> Square Feet, lf -> Linear Feet.
  7. Package tokens found in Size (FG, cont/container, B&B) are listed in Package, "/"-joined.

Two kinds of row are protected from the item steps: rows that held a header at step 1
(matched, flagged or coloured — their label, note and cells are left as they are), and rows
that had BOTH a Description and a Size to begin with (the spec says that never happens on a
well-formed sheet; when it does, the row is red-filled and left exactly as it came).

Pure openpyxl (no pandas). Handles .xlsx and .xlsm — a macro workbook is loaded with
keep_vba=True so its VBA, tables (incl. calculated columns), formulas and defined names are
carried through to the output untouched.
"""

from __future__ import annotations

import datetime as _dt
import difflib
import re
import warnings
import zipfile
from copy import copy
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

# openpyxl grumbles about extension parts it doesn't model (the Landscape Hub template has a
# few). Nothing we rely on lives there and the warning would just pollute the tool's stderr.
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

GROUPINGS_FILE = Path(__file__).parent / "takeoff-groupings.txt"
SUPPORTED_EXTENSIONS = (".xlsx", ".xlsm")

# "Flag the entire cell by coloring it red" — a solid fill, not red text: red *text* already
# means something on these sheets (the Landscape / IRRIGATION section markers), and a filled
# cell is unmistakable in a column of white ones.
RED_FILL = PatternFill(fill_type="solid", start_color="FFFF0000", end_color="FFFF0000")

# Header names as they appear on the worksheets (case-insensitive). The row that carries a
# Description AND a Size header is the header row; everything else is optional.
HEADER_ALIASES = {
    "description": ("description", "desc", "item", "items", "item description"),
    "size": ("size",),
    "package": ("package", "packages", "pkg"),
    "notes": ("notes", "note", "comments", "comment"),
    "uom": ("uom", "unit", "units", "unit of measure", "u/m"),
    "qty": ("qty 1", "qty1", "qty", "quantity"),
}
REQUIRED_COLUMNS = ("description", "size")
MAX_HEADER_SCAN_ROWS = 60

UOM_MAP = {
    "count": "Unit", "ea": "Unit", "each": "Unit",
    "sf": "Square Feet",
    "lf": "Linear Feet",
}

# Package tokens to lift out of Size, in canonical form. Word-bounded so "cont" never fires on
# "continuous"/"contractor" and "fg" never on the "G" of "45 G". Spelled-out forms of the same
# packages ("field grown", "balled & burlapped") are accepted too.
PACKAGE_PATTERNS = (
    (re.compile(r"\bf\.?g\.?(?![a-z])|\bfield[\s-]?grown\b", re.I), "FG"),
    (re.compile(r"\bcont(?:ainers?)?\.?(?![a-z])", re.I), "Container"),
    (re.compile(r"\bb\.?\s*(?:&|and|\+)\s*b\.?s?(?![a-z])|\bball(?:ed)?\s*(?:&|and)\s*burlap(?:ped)?\b", re.I), "B&B"),
)

SOIL_MULCH_RE = re.compile(r"\b(soil|topsoil|mulch)\b", re.I)
DEPTH_RE = re.compile(r"\bdepth\b", re.I)

# A measurement: number + unit, in the spellings these sheets use. Three shapes, longest first:
# feet-and-inches ("1'-6\""), a range ("3-4\"", "3\"-4\"", "2\" to 3\"", "6 in. to 8 in."), a
# single value ("3\"", "3 in.", "3-inch"). Numbers may be decimals, fractions or mixed numbers
# ("2 1/2", "1-1/2", "1½"); the lookbehind stops a match starting inside "1/2" at the "2".
_NUM = r"(?<![\d/.])(?:\d+[-\s]+\d+/\d+|\d+/\d+|\d+(?:\.\d+)?|\d*[½¼¾⅛⅜⅝⅞])"
_IN = r"(?:[\"”″]|''|in(?:ch(?:es)?)?\.?(?![a-z]))"
_FT = r"(?:['’′]|ft\.?(?![a-z])|feet\b|foot\b)"
_UNIT = rf"(?:{_IN}|{_FT}|mm\b|cm\b)"
_SEP = r"\s*-?\s*"                     # "3in", "3 in", "3-inch"
MEASURE_RE = re.compile(
    rf"{_NUM}\s*{_FT}\s*-?\s*{_NUM}\s*{_IN}"                                # 1'-6"
    rf"|{_NUM}(?:{_SEP}{_UNIT})?\s*(?:-|–|to)\s*{_NUM}{_SEP}{_UNIT}"        # 3-4"   3"-4"   2" to 3"
    rf"|{_NUM}{_SEP}{_UNIT}",                                                # 3"   3 in.   3-inch
    re.I,
)

# --- group matching ----------------------------------------------------------------------
_STOPWORDS = {"and", "or", "the", "of", "for", "a", "an", "to", "in", "with"}

# Estimator vocabulary that means one of the groupings' own words. Applied to the header text
# before tokenising, so "Ground Cover" meets "groundcover" exactly and "Hydroseed" meets
# "seed". The last two are Arazoza's own OST group names, learnt from their finished Vela Cove
# worksheet: "Bed Areas" is where the mulch/rock bed materials live, "Edging" is bed prep.
# Extend HERE (not takeoff-groupings.txt, which is the user's list of labels). Abbreviations
# like a bare "GC" are deliberately NOT expanded: Arazoza files "Shrubs & Gc" under Shrubs.
ALIASES = (
    (r"\bground[\s-]?covers?\b", "groundcover"),
    (r"\bhydro[\s-]?seed(?:ing|ed)?\b", "seed"),
    (r"\bsodding\b", "sod"),
    (r"\bhedges?\b", "shrub"),
    (r"\bpavers?\b", "hardscape"),
    (r"\brip[\s-]?rap\b", "rock"),
    (r"\btop[\s-]?soil\b", "soil"),
    (r"\bplanting (?:soil )?mix\b", "soil"),
    (r"\bpine[\s-]?straw\b", "mulch"),
    (r"\btransplant(?:s|ed|ing)?\b", "relocate"),
    (r"\bbed\s+areas?\b", "mulch"),
    (r"\b(?:bed\s+)?edging\b", "bed preparation"),
)
_ALIASES = tuple((re.compile(p, re.I), rep) for p, rep in ALIASES)

_MATCH_EXACT = 1.0        # the same word (after lower-casing / de-pluralising)
_MATCH_INFLECTED = 0.9    # the same stem: protect/protection, relocate/relocations, mulch/mulching
_MATCH_WEAK = 0.7         # a resemblance only: prep/preparation, street/tree, project/protect
_STRONG = _MATCH_INFLECTED
_SUFFIXES = ("ations", "ation", "ments", "ment", "ings", "ing", "ions", "ion", "ers", "ies", "es", "ed", "er", "s", "e")


class FormatterError(Exception):
    """Something about the input makes formatting impossible or unsafe. The message is written
    for the person who sent the file, not for a developer."""


# ---------------------------------------------------------------------------------------
# Small helpers
# ---------------------------------------------------------------------------------------

def _txt(value) -> str:
    """Cell value as stripped text ('' for None). Dates show as their date, not 'YYYY-MM-DD 00:00:00'."""
    if value is None:
        return ""
    if isinstance(value, _dt.datetime) and value.time() == _dt.time(0):
        return value.date().isoformat()
    return str(value).strip()


def _empty(value) -> bool:
    return _txt(value) == ""


def _is_formula(value) -> bool:
    return isinstance(value, str) and value.startswith("=")


def is_black_font(cell) -> bool:
    """Is this cell's text the default black? Anything else (red section markers, blue notes,
    grey guidance text) is left alone by the group matcher, by the spec.

    openpyxl reports a font colour four ways: absent (default = black), an ARGB string, a
    theme index (1 = "Text 1" = black in Office themes; 0 is the *background*), or a legacy
    palette index (8 = black, 64 = "automatic", which renders black)."""
    color = cell.font.color if cell.font is not None else None
    if color is None:
        return True
    ctype = getattr(color, "type", None)
    if ctype == "rgb":
        rgb = str(color.rgb).upper()
        return rgb.endswith("000000")            # FF000000 / 00000000, alpha ignored
    if ctype == "theme":
        return color.theme == 1 and abs(color.tint or 0.0) < 0.05
    if ctype == "indexed":
        return color.indexed in (8, 64)
    if ctype == "auto":
        return True
    return False


def _norm_header(value) -> str:
    return re.sub(r"\s+", " ", _txt(value).lower()).rstrip(":").strip()


def _singular(word: str) -> str:
    if len(word) > 3 and word.endswith("s") and not word.endswith("ss"):
        return word[:-1]
    return word


def _stem(word: str) -> str:
    """Crude two-pass suffix stripper, applied identically to both sides of a comparison so
    'relocation' and 'relocate' both become 'relocat'. Never shortens a word below 4 letters."""
    for _ in range(2):
        for suffix in _SUFFIXES:
            if word.endswith(suffix) and len(word) - len(suffix) >= 4:
                word = word[: -len(suffix)]
                break
    return word


def dealias(text: str) -> str:
    """Estimator vocabulary rewritten to the groupings' own words ("Pine Straw" -> "mulch").
    Used by the matcher and by the soil/mulch test, so both know the same synonyms."""
    t = text.lower()
    for pattern, replacement in _ALIASES:
        t = pattern.sub(replacement, t)
    return t


def _tokens(text: str) -> list[str]:
    """Words that carry meaning for group matching: aliases resolved, lower-cased, '&' -> 'and'
    (then dropped as a stopword), parentheticals like "(SF)" removed, a leading "1 - " /
    "01." code removed, crude singularisation so "Trees" meets "tree"."""
    t = text.lower().replace("&", " and ")
    t = re.sub(r"\(.*?\)", " ", t)
    t = re.sub(r"^\s*\d+[\s.\-:)]*", " ", t)
    t = dealias(t)
    t = re.sub(r"[^a-z0-9]+", " ", t)
    return [_singular(w) for w in t.split() if w not in _STOPWORDS]


def _token_match(a: str, b: str) -> float:
    """1.0 for the same word, 0.9 for the same stem, 0.7 for a mere resemblance, 0 otherwise.
    Only the first two count as evidence that a header names a grouping; the third exists so
    "Bed Prep" outscores a bare "Bed" — it can never make a match on its own."""
    if a == b:
        return _MATCH_EXACT
    sa, sb = _stem(a), _stem(b)
    if sa == sb or (min(len(sa), len(sb)) >= 5 and (sa.startswith(sb) or sb.startswith(sa))):
        return _MATCH_INFLECTED          # relocation/relocate -> 'reloc'/'relocat'
    short, long_ = sorted((a, b), key=len)
    if len(short) >= 4 and short in long_:
        return _MATCH_WEAK
    prefix = 0
    for x, y in zip(a, b):
        if x != y:
            break
        prefix += 1
    if prefix >= 5:
        return _MATCH_WEAK
    if difflib.SequenceMatcher(None, a, b).ratio() >= 0.85:
        return _MATCH_WEAK
    return 0.0


# ---------------------------------------------------------------------------------------
# Takeoff groupings
# ---------------------------------------------------------------------------------------

@dataclass(frozen=True)
class Grouping:
    label: str                 # exactly what gets written into the cell, e.g. "1 - Trees & Palms"
    name: str                  # "Trees & Palms"
    tokens: tuple              # ("tree", "palm")


def load_groupings(path: Path = GROUPINGS_FILE) -> list[Grouping]:
    groupings: list[Grouping] = []
    for raw in Path(path).read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        name = line.split(" - ", 1)[1] if " - " in line else line
        toks = tuple(_tokens(name))
        if not toks:
            continue
        groupings.append(Grouping(label=line, name=name.strip(), tokens=toks))
    if not groupings:
        raise FormatterError(f"The takeoff groupings list at {path} is empty.")
    return groupings


@dataclass
class GroupMatch:
    label: str | None          # the grouping to write, or None when flagged
    reason: str                # "matched" | "no match" | "ambiguous: A / B"
    score: float = 0.0
    candidates: tuple = ()     # labels still in the running when the header alone couldn't decide


def match_grouping(text: str, groupings: list[Grouping]) -> GroupMatch:
    """Closest grouping for a header cell's text — or None with the reason it must be flagged.

    A grouping is a candidate only if at least one word of the header IS one of its words
    (exactly, or the same stem: "Mulching" ~ "mulch"). Resemblances ("Streetscape" ~ "tree",
    "Project" ~ "protect") never qualify on their own, so they can't quietly relabel a header.

    Among candidates, one that explains every strongly-matched word another one explains
    wins ("Remove & Replace" -> Remove/Replace, not Replacements). Two candidates that each
    explain words the other doesn't are exactly the spec's "more than one match" — "Shrubs &
    Groundcover", "Trees, Palms & Shrubs", "Tree Protection" — and the cell is flagged."""
    cell_tokens = _tokens(text)
    if not cell_tokens:
        return GroupMatch(None, "no match")

    candidates = []                      # (grouping, strong-token set, score)
    for g in groupings:
        strong: set = set()
        score = 0.0
        for ct in cell_tokens:
            best = max((_token_match(ct, gt) for gt in g.tokens), default=0.0)
            score += best
            if best >= _STRONG:
                strong.add(ct)
        if strong:
            candidates.append((g, strong, score))
    if not candidates:
        return GroupMatch(None, "no match")

    # Drop anything another candidate strictly dominates (explains a superset of its words).
    top = [c for c in candidates if not any(o[1] > c[1] for o in candidates)]
    if len(top) == 1:
        g, _, score = top[0]
        return GroupMatch(g.label, "matched", score)
    if len({frozenset(c[1]) for c in top}) == 1:
        # Same words, different groupings ("Replacements" vs Remove/Replace): exact beats stem.
        top.sort(key=lambda c: -c[2])
        if top[0][2] - top[1][2] >= 0.05:
            g, _, score = top[0]
            return GroupMatch(g.label, "matched", score)
    return GroupMatch(None, "ambiguous: " + " / ".join(c[0].label for c in top), max(c[2] for c in top),
                      candidates=tuple(c[0].label for c in top))


def resolve_with_items(header: GroupMatch, item_texts: list[str], groupings: list[Grouping]) -> GroupMatch:
    """Second opinion for a header the text alone couldn't place: what do the items beneath it
    say? This is what the estimator does by hand — "07 - Seed & Sod" holding nothing but sod
    rows is the Sod group. Each item is matched like a header; items that name exactly one
    grouping vote for it.

    - Ambiguous header: if exactly one of its candidates gets votes, that one wins.
    - Unmatched header: if every voting item agrees on one grouping, at least two items voted
      and they are at least half the items, that grouping wins.
    Anything less stays as it was (and gets flagged red)."""
    if header.label is not None or not item_texts:
        return header
    votes: dict = {}
    for text in item_texts:
        m = match_grouping(text, groupings)
        if m.label is not None:
            votes[m.label] = votes.get(m.label, 0) + 1
    if not votes:
        return header
    if header.candidates:
        backed = [c for c in header.candidates if votes.get(c)]
        if len(backed) == 1:
            return GroupMatch(backed[0], "matched by its items", header.score)
        return header
    if len(votes) == 1:
        (label, n), = votes.items()
        if n >= 2 and n * 2 >= len(item_texts):
            return GroupMatch(label, "matched by its items", 0.0)
    return header


# ---------------------------------------------------------------------------------------
# Locating the table on the sheet
# ---------------------------------------------------------------------------------------

@dataclass
class Layout:
    ws: object
    header_row: int
    cols: dict                 # role -> column index (1-based); only roles that exist
    first_row: int
    last_row: int
    table_end: int | None = None       # last row of the Excel table the header belongs to, if any
    beyond_table: tuple = ()           # data rows found below that table

    def col(self, role: str) -> int | None:
        return self.cols.get(role)

    def cell(self, row: int, role: str):
        return self.ws.cell(row=row, column=self.cols[role])

    def rows(self):
        return range(self.first_row, self.last_row + 1)


def _header_map(ws, row: int) -> dict:
    """role -> column for the recognised headers in this row (leftmost wins per role, so the
    first 'UOM' is the primary one and 'UOM2' further right is not)."""
    found: dict = {}
    for cell in ws[row]:
        h = _norm_header(cell.value)
        if not h:
            continue
        for role, aliases in HEADER_ALIASES.items():
            if h in aliases and role not in found:
                found[role] = cell.column
    return found


def _has_data(ws, row: int, cols: dict) -> bool:
    return any(not _empty(ws.cell(row, c).value) for c in cols.values())


def _data_row_count(ws, header_row: int, cols: dict) -> int:
    return sum(1 for r in range(header_row + 1, ws.max_row + 1) if _has_data(ws, r, cols))


def _table_end(ws, header_row: int, desc_col: int) -> int | None:
    """If the header row is the header of an Excel Table, that table's last row."""
    for ref in ws.tables.values():
        try:
            min_col, min_row, max_col, max_row = range_boundaries(ref if isinstance(ref, str) else ref.ref)
        except Exception:
            continue
        if min_row == header_row and min_col <= desc_col <= max_col:
            return max_row
    return None


def find_layout(wb, sheet_name: str | None = None) -> Layout:
    """Find the sheet + header row holding the worksheet columns. With no `sheet_name`, every
    sheet is considered and the visible one with the most data rows wins (the Landscape Hub
    template carries hidden 'Project Break-Out' sheets with the same headers and no data).

    The data region runs to the end of the Excel table the header belongs to — plus any rows
    with data pasted below it, which are formatted too and reported, since the table's import
    formulas won't cover them until someone extends the table in Excel."""
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise FormatterError(
                f"There's no sheet called {sheet_name!r} in this workbook. "
                f"Sheets: {', '.join(wb.sheetnames)}."
            )
        if not isinstance(wb[sheet_name], Worksheet):
            raise FormatterError(f"{sheet_name!r} is a chart sheet, not a worksheet with cells.")
        candidates = [wb[sheet_name]]
    else:
        candidates = [ws for ws in wb.worksheets if isinstance(ws, Worksheet)]

    best: tuple | None = None
    for ws in candidates:
        for row in range(1, min(ws.max_row, MAX_HEADER_SCAN_ROWS) + 1):
            cols = _header_map(ws, row)
            if all(role in cols for role in REQUIRED_COLUMNS):
                visible = 1 if ws.sheet_state == "visible" else 0
                key = (visible, _data_row_count(ws, row, cols))
                if best is None or key > best[0]:
                    best = (key, ws, row, cols)
                break       # first header row per sheet

    if best is None:
        looked = ", ".join(ws.title for ws in candidates)
        raise FormatterError(
            "I couldn't find the worksheet table: no row has both a 'Description' and a 'Size' "
            f"header (looked on: {looked}). If the headers are named differently, tell me which "
            "sheet and columns to use."
        )
    _, ws, header_row, cols = best
    table_end = _table_end(ws, header_row, cols["description"])
    last = max(table_end or ws.max_row, header_row)
    beyond = ()
    if table_end is not None and ws.max_row > table_end:
        beyond = tuple(r for r in range(table_end + 1, ws.max_row + 1) if _has_data(ws, r, cols))
        if beyond:
            last = beyond[-1]
    return Layout(ws=ws, header_row=header_row, cols=cols, first_row=header_row + 1, last_row=last,
                  table_end=table_end, beyond_table=beyond)


# ---------------------------------------------------------------------------------------
# The report
# ---------------------------------------------------------------------------------------

@dataclass
class Report:
    sheet: str = ""
    header_row: int = 0
    columns: dict = field(default_factory=dict)          # role -> column letter
    # step 1
    groups_matched: list = field(default_factory=list)    # (row, from, to)
    groups_by_items: list = field(default_factory=list)   # (row, from, to)  subset of the above, placed by the items beneath
    groups_unchanged: list = field(default_factory=list)  # (row, label)   already in target form
    groups_flagged: list = field(default_factory=list)    # (row, text, reason)
    groups_skipped_colored: list = field(default_factory=list)   # (row, text)
    # step 2
    items_moved: int = 0
    conflicts: list = field(default_factory=list)         # (row, description, size)
    # step 3
    notes_moved: int = 0
    notes_blocked: list = field(default_factory=list)     # (row, size, notes)
    # step 4
    merged: int = 0
    merge_skipped: list = field(default_factory=list)     # (row, description, size)  size already the suffix
    # step 5
    depth_rows: list = field(default_factory=list)        # (row, description, size, package)
    depth_unparsed: list = field(default_factory=list)    # (row, description)
    # step 6
    uom_converted: dict = field(default_factory=dict)     # target -> {source spelling -> count}
    # step 7
    packages_set: list = field(default_factory=list)      # (row, package)
    # anything odd
    warnings: list = field(default_factory=list)

    @property
    def flagged_count(self) -> int:
        return len(self.groups_flagged) + len(self.conflicts)

    def summary(self, input_name: str, output_name: str) -> str:
        """What the bot tells the user. Written to be relayed nearly verbatim, so it leads with
        what needs a human (red cells) and keeps the counts to one line each."""
        lines = [f"Formatted sheet '{self.sheet}' of {input_name} → {output_name}."]

        if self.groups_flagged or self.conflicts:
            n = self.flagged_count
            lines.append(f"{n} cell{'s' if n != 1 else ''} filled RED for a human to check:")
            for row, text, reason in self.groups_flagged[:20]:
                lines.append(f"  - row {row}: '{_shorten(text)}' ({reason})")
            for row, desc, size in self.conflicts[:10]:
                lines.append(f"  - row {row}: Description '{_shorten(desc)}' and Size '{_shorten(size)}' were BOTH "
                             f"filled, so that row was left exactly as it came")
            hidden = n - min(len(self.groups_flagged), 20) - min(len(self.conflicts), 10)
            if hidden > 0:
                lines.append(f"  - …and {hidden} more (see the red cells in the file)")
        else:
            lines.append("No cells needed flagging.")

        matched = self.groups_matched
        if matched or self.groups_unchanged:
            pairs = ", ".join(f"'{_shorten(a, 24)}' → '{b}'" for _, a, b in _dedupe_pairs(matched)[:12])
            extra = f" ({len(self.groups_unchanged)} already in final form)" if self.groups_unchanged else ""
            lines.append(f"Groups: {len(matched)} header{'s' if len(matched) != 1 else ''} relabelled{extra}"
                         + (f": {pairs}" if pairs else "") + ".")
            if self.groups_by_items:
                by = ", ".join(f"'{_shorten(a, 24)}' → '{b}'" for _, a, b in self.groups_by_items[:6])
                lines.append(f"Of those, {len(self.groups_by_items)} named more than one group (or none) and "
                             f"{'was' if len(self.groups_by_items) == 1 else 'were'} placed by the items listed "
                             f"under {'it' if len(self.groups_by_items) == 1 else 'them'}: {by}. Worth a glance.")
        if self.groups_skipped_colored:
            n = len(self.groups_skipped_colored)
            lines.append(f"Skipped {n} coloured (non-black) Description cell{'s' if n != 1 else ''} as "
                         f"section marker{'s' if n != 1 else ''} (e.g. '{_shorten(self.groups_skipped_colored[0][1], 24)}').")

        lines.append(
            f"Rows: {self.items_moved} item{'s' if self.items_moved != 1 else ''} moved Size → Description; "
            f"{self.notes_moved} Notes moved → Size; {self.merged} Description cells merged with ' - ' + Size"
            + (f" ({len(self.merge_skipped)} already ended with their size: row"
               f"{'s' if len(self.merge_skipped) != 1 else ''} {', '.join(str(r) for r, _, _ in self.merge_skipped[:8])})"
               if self.merge_skipped else "") + "."
        )
        if self.depth_rows:
            ex = "; ".join(f"row {r}: Size '{s}', Package '{p}'" for r, _, s, p in self.depth_rows[:4])
            lines.append(f"Soil/mulch depth: {len(self.depth_rows)} row{'s' if len(self.depth_rows) != 1 else ''} "
                         f"set ({ex}{'; …' if len(self.depth_rows) > 4 else ''}).")
        if self.uom_converted:
            parts = []
            for target, sources in self.uom_converted.items():
                total = sum(sources.values())
                spellings = ", ".join(sorted(sources, key=str.lower))
                parts.append(f"{total} → {target} (from {spellings})")
            lines.append("UOM converted: " + "; ".join(parts) + ".")
        if self.packages_set:
            counts: dict = {}
            for _, p in self.packages_set:
                counts[p] = counts.get(p, 0) + 1
            lines.append("Packages filled from Size on "
                         f"{len(self.packages_set)} row{'s' if len(self.packages_set) != 1 else ''}: "
                         + ", ".join(f"{k} ×{v}" for k, v in sorted(counts.items())) + ".")
        for w in self.depth_unparsed[:5]:
            lines.append(f"Note: row {w[0]} mentions 'depth' but I couldn't read a measurement from "
                         f"'{_shorten(w[1])}', so its Size was left alone.")
        for w in self.notes_blocked[:5]:
            lines.append(f"Note: row {w[0]} already had Size '{_shorten(w[1])}', so its Notes "
                         f"'{_shorten(w[2])}' stayed put.")
        lines.extend(f"Note: {w}" for w in self.warnings[:6])
        held_back = (max(0, len(self.depth_unparsed) - 5) + max(0, len(self.notes_blocked) - 5)
                     + max(0, len(self.warnings) - 6))
        if held_back:
            lines.append(f"(…and {held_back} more note{'s' if held_back != 1 else ''} I've left out to keep this short.)")
        return "\n".join(lines)


def _shorten(text, n: int = 60) -> str:
    t = _txt(text).replace("\n", " ")
    return t if len(t) <= n else t[: n - 1] + "…"


def _dedupe_pairs(pairs):
    seen, out = set(), []
    for row, a, b in pairs:
        key = (a.strip().lower(), b)
        if key not in seen:
            seen.add(key)
            out.append((row, a, b))
    return out


# ---------------------------------------------------------------------------------------
# The steps — only ever called, in order, by format_worksheet
# ---------------------------------------------------------------------------------------

def _set(cell, value, report: Report) -> bool:
    """Write a value, unless the target is part of a merged range (openpyxl refuses)."""
    if isinstance(cell, MergedCell):
        report.warnings.append(f"{cell.coordinate} is inside a merged range; I couldn't write to it.")
        return False
    cell.value = value
    return True


def _move(src, dst, report: Report) -> bool:
    """Move a value (and its font, so red/blue emphasis travels with it) from src to dst."""
    if not _set(dst, src.value, report):
        return False
    try:
        dst.font = copy(src.font)
    except Exception:
        pass
    src.value = None
    return True


def _find_conflicts(layout: Layout) -> set:
    """Rows with BOTH a Description and a Size value. The spec says this never happens on a
    well-formed sheet (item names live in Size, headers in Description, never both on one
    row), so when it does the row is left exactly as it came — red-filled, untouched by every
    step — for a person to sort out. Reported by step 2, whose precondition it breaks."""
    return {
        r for r in layout.rows()
        if not _empty(layout.cell(r, "description").value) and not _empty(layout.cell(r, "size").value)
    }


def _step1_match_groups(layout: Layout, groupings: list[Grouping], report: Report, skip: set) -> dict:
    """Relabel black Description headers; red-fill the unmatched/ambiguous. A header the text
    alone can't place gets a second opinion from the items beneath it (resolve_with_items).

    Returns row -> label for EVERY row that held a header: the grouping label where one
    matched (or was already there), None where the header was flagged or is a coloured
    section marker. Later steps use the keys to leave header rows alone, and the depth rule
    uses the values to know which group — if any recognised one — an item sits under."""
    header_rows: dict = {}
    # A row already holding text in Description is a group header — unless it also carries a
    # quantity, which no header on these sheets ever does. That one is an item somebody typed
    # into the right column already; relabelling it with a grouping would destroy its name.
    qty_col = layout.col("qty")
    candidates = [
        r for r in layout.rows()
        if r not in skip and not _empty(layout.cell(r, "description").value)
        and (qty_col is None or _empty(layout.cell(r, "qty").value))
    ]
    for i, r in enumerate(candidates):
        cell = layout.cell(r, "description")
        text = _txt(cell.value)
        header_rows[r] = None
        if not is_black_font(cell):
            report.groups_skipped_colored.append((r, text))
            continue
        m = match_grouping(text, groupings)
        if m.label is None:
            # The items under this header (still sitting in Size at this point) get a say.
            until = candidates[i + 1] if i + 1 < len(candidates) else layout.last_row + 1
            items = [_txt(layout.cell(k, "size").value) for k in range(r + 1, until)]
            m = resolve_with_items(m, [t for t in items if t], groupings)
        if m.label is None:
            if not isinstance(cell, MergedCell):
                cell.fill = RED_FILL
            report.groups_flagged.append((r, text, m.reason))
            continue
        header_rows[r] = m.label
        if text == m.label:
            report.groups_unchanged.append((r, m.label))
        else:
            _set(cell, m.label, report)
            report.groups_matched.append((r, text, m.label))
            if m.reason == "matched by its items":
                report.groups_by_items.append((r, text, m.label))
    return header_rows


def _step2_size_to_description(layout: Layout, report: Report, conflicts: set) -> set:
    """Every row with a Size value: Size -> Description. Description must be empty on those
    rows (it always is on a well-formed sheet); when it isn't, nothing moves and the
    Description cell is red-filled so the clash is visible."""
    moved: set = set()
    for r in layout.rows():
        size = layout.cell(r, "size")
        if _empty(size.value):
            continue
        desc = layout.cell(r, "description")
        if r in conflicts:
            if not isinstance(desc, MergedCell):
                desc.fill = RED_FILL
            report.conflicts.append((r, _txt(desc.value), _txt(size.value)))
            continue
        if _move(size, desc, report):
            moved.add(r)
            report.items_moved += 1
    return moved


def _step3_notes_to_size(layout: Layout, report: Report, skip: set) -> None:
    if layout.col("notes") is None:
        report.warnings.append("There is no 'Notes' column on this sheet, so there was no sizing to move into 'Size'.")
        return
    for r in layout.rows():
        if r in skip:
            continue
        notes = layout.cell(r, "notes")
        if _empty(notes.value):
            continue
        size = layout.cell(r, "size")
        if not _empty(size.value):
            report.notes_blocked.append((r, _txt(size.value), _txt(notes.value)))
            continue
        if isinstance(notes.value, (_dt.datetime, _dt.date)):
            report.warnings.append(
                f"row {r}: Notes held a date ({_txt(notes.value)}) — Excel may have turned a size like "
                f"'3-4' into a date; please check that cell.")
        if _move(notes, size, report):
            report.notes_moved += 1


def _step4_merge_size_into_description(layout: Layout, report: Report, skip: set) -> None:
    for r in layout.rows():
        if r in skip:
            continue
        desc = layout.cell(r, "description")
        size_raw = layout.cell(r, "size").value
        size = _txt(size_raw)
        d = _txt(desc.value)
        if not d or not size:
            continue
        if _is_formula(desc.value) or _is_formula(size_raw):
            report.warnings.append(f"row {r}: Description/Size holds a formula, so I didn't merge them.")
            continue
        if d.lower() == size.lower() or d.lower().endswith(" - " + size.lower()):
            report.merge_skipped.append((r, d, size))
            continue
        if _set(desc, f"{d} - {size}", report):
            report.merged += 1


def _group_for_row(row: int, header_rows: dict) -> str | None:
    """The label of the nearest header above this row — None if that header was flagged or is
    a coloured section marker (so a Soil group never leaks past an 'IRRIGATION' marker), or
    if there is no header above at all."""
    label = None
    for hr in sorted(header_rows):
        if hr < row:
            label = header_rows[hr]
        else:
            break
    return label


def _is_soil_or_mulch(description: str, group_label: str | None) -> bool:
    """Is this a soil or mulch item? Its own words decide first — through the same alias list
    the matcher uses, so "Pine Straw 3\" depth" and "Planting Mix 6\" depth" count as mulch and
    soil the way an estimator reads them — then the group it sits under."""
    if SOIL_MULCH_RE.search(dealias(description)):
        return True
    return bool(group_label) and bool(SOIL_MULCH_RE.search(group_label))


def extract_depth(text: str) -> str | None:
    """The measurement that goes with the word 'depth' in an item description: '3"' from
    'Mulch 3" Depth' or '3" Depth of Mulch', '12"' from 'Planting Soil Mix 12" Depth',
    '3"-4"' from 'Mulch 3"-4" depth', '2 1/2"' from 'Mulch 2 1/2" depth'. When several
    measurements appear, the one nearest the word wins. None if there's no 'depth' or no
    measurement at all."""
    depth = DEPTH_RE.search(text)
    if not depth:
        return None
    measures = list(MEASURE_RE.finditer(text))
    if not measures:
        return None
    best = min(measures, key=lambda m: min(abs(m.start() - depth.end()), abs(depth.start() - m.end())))
    return re.sub(r"\s+", " ", best.group(0)).strip()


def _add_package(existing, addition: str) -> str:
    parts = [p.strip() for p in _txt(existing).split("/") if p.strip() and p.strip() != "-"]
    if addition.lower() not in {p.lower() for p in parts}:
        parts.append(addition)
    return "/".join(parts)


def _step5_depth(layout: Layout, header_rows: dict, report: Report, skip: set) -> None:
    for r in layout.rows():
        if r in skip:
            continue
        desc = _txt(layout.cell(r, "description").value)
        if not desc or not DEPTH_RE.search(desc):
            continue
        if not _is_soil_or_mulch(desc, _group_for_row(r, header_rows)):
            continue
        depth = extract_depth(desc)
        if depth is None:
            report.depth_unparsed.append((r, desc))
            continue
        size_cell = layout.cell(r, "size")
        if not _set(size_cell, depth, report):
            continue
        package = None
        if layout.col("package") is None:
            # Say so rather than reporting a depth row with a blank Package, which reads as if
            # the write had happened.
            report.warnings.append(
                f"row {r}: Size is now the depth ({depth}), but this sheet has no 'Package' column "
                f"to put 'Depth' in.")
        else:
            pkg = layout.cell(r, "package")
            package = _add_package(pkg.value, "Depth")
            _set(pkg, package, report)
        report.depth_rows.append((r, desc, depth, package or "—"))


def _step6_uom(layout: Layout, report: Report, skip: set) -> None:
    if layout.col("uom") is None:
        report.warnings.append("There is no 'UOM' column on this sheet, so no units were converted.")
        return
    for r in layout.rows():
        if r in skip:
            continue
        cell = layout.cell(r, "uom")
        raw = _txt(cell.value)
        if not raw:
            continue
        key = re.sub(r"\s+", " ", raw.lower().replace(".", "")).strip()
        new = UOM_MAP.get(key)
        if new is None or new == raw:
            continue
        if _set(cell, new, report):
            sources = report.uom_converted.setdefault(new, {})
            sources[raw] = sources.get(raw, 0) + 1


def find_packages(size_text: str) -> list[str]:
    """Canonical package names mentioned in a Size string, in order of appearance, deduped."""
    hits = []
    for pattern, canonical in PACKAGE_PATTERNS:
        for m in pattern.finditer(size_text):
            hits.append((m.start(), canonical))
    out: list[str] = []
    for _, canonical in sorted(hits):
        if canonical not in out:
            out.append(canonical)
    return out


def _step7_packages(layout: Layout, report: Report, skip: set) -> None:
    if layout.col("package") is None:
        report.warnings.append("There is no 'Package' column on this sheet, so package types were not copied out of Size.")
        return
    for r in layout.rows():
        if r in skip:
            continue
        size = _txt(layout.cell(r, "size").value)
        if not size:
            continue
        found = find_packages(size)
        if not found:
            continue
        pkg = layout.cell(r, "package")
        before = _txt(pkg.value)
        after = before
        for name in found:
            after = _add_package(after, name)
        if after != before and _set(pkg, after, report):
            report.packages_set.append((r, after))


# ---------------------------------------------------------------------------------------
# Entry points
# ---------------------------------------------------------------------------------------

def _preflight(layout: Layout, force: bool) -> None:
    """Refuse a sheet that isn't in the raw Arazoza layout. The tell is step 2's precondition:
    a raw sheet has item names sitting in Size with Description empty on those rows. A sheet
    with nothing to move is either already formatted or laid out some other way — and running
    step 1 on it would paint every item name red, which is the one outcome worse than doing
    nothing."""
    movable = sum(
        1 for r in layout.rows()
        if not _empty(layout.cell(r, "size").value) and _empty(layout.cell(r, "description").value)
    )
    if movable or force:
        return
    raise FormatterError(
        f"Sheet '{layout.ws.title}' doesn't look like a raw Arazoza worksheet: every item already "
        "sits in 'Description' and there is nothing in 'Size' to move over. If this file has "
        "already been formatted there's nothing to do; if its columns are laid out differently, "
        "tell me which sheet/columns to use. (Do NOT retry with force=true on your own — that is "
        "the user's call, and it would flag every item red.)"
    )


def format_worksheet(wb, sheet_name: str | None = None, force: bool = False,
                     groupings: list[Grouping] | None = None) -> Report:
    """Run the seven steps, in order, on an open workbook. Mutates `wb`; returns the Report."""
    groupings = groupings or load_groupings()
    layout = find_layout(wb, sheet_name)
    _preflight(layout, force)

    report = Report(
        sheet=layout.ws.title,
        header_row=layout.header_row,
        columns={role: layout.ws.cell(layout.header_row, c).column_letter for role, c in layout.cols.items()},
    )
    if layout.beyond_table:
        rows = layout.beyond_table
        span = f"row {rows[0]}" if len(rows) == 1 else f"rows {rows[0]}–{rows[-1]}"
        report.warnings.append(
            f"{span} sit below the sheet's Excel table (which ends at row {layout.table_end}). I formatted "
            f"them too, but the table — and the import formulas that live in it — won't cover them until "
            f"the table is extended in Excel.")

    conflicts = _find_conflicts(layout)
    # The order below is the specification. Do not reorder.
    header_rows = _step1_match_groups(layout, groupings, report, skip=conflicts)
    _step2_size_to_description(layout, report, conflicts)
    # Header rows (matched, flagged or coloured) are done once step 1 has seen them: a stray
    # Notes value on one must not be glued onto the label, and the depth/package rules are
    # about items. Conflict rows stay exactly as they came.
    frozen = conflicts | set(header_rows)
    _step3_notes_to_size(layout, report, skip=frozen)
    _step4_merge_size_into_description(layout, report, skip=frozen)
    _step5_depth(layout, header_rows, report, skip=frozen)
    _step6_uom(layout, report, skip=conflicts)
    _step7_packages(layout, report, skip=frozen)
    return report


def output_name_for(input_path: str | Path) -> str:
    p = Path(input_path)
    return f"{p.stem} - formatted{p.suffix.lower()}"


def _has_drawings(path: Path) -> bool:
    try:
        with zipfile.ZipFile(path) as z:
            return any(n.startswith("xl/drawings/drawing") for n in z.namelist())
    except Exception:
        return False


def format_file(input_path: str | Path, output_path: str | Path, sheet_name: str | None = None,
                force: bool = False) -> Report:
    """Load, format, save. The original is never written to."""
    input_path, output_path = Path(input_path), Path(output_path)
    ext = input_path.suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        raise FormatterError(
            f"I can only format .xlsx or .xlsm workbooks, and this is {ext or 'an extension-less file'!r}. "
            "If it's an old .xls, please re-save it as .xlsx and send that."
        )
    keep_vba = ext == ".xlsm"
    try:
        wb = openpyxl.load_workbook(input_path, keep_vba=keep_vba)
    except Exception as err:
        raise FormatterError(f"I couldn't open that workbook ({type(err).__name__}: {err}). Is it a valid Excel file?") from err
    report = format_worksheet(wb, sheet_name=sheet_name, force=force)
    if keep_vba and _has_drawings(input_path):
        report.warnings.append(
            "this is a macro workbook: its cells, formulas, tables, names and macros are all carried over, "
            "but shapes or buttons drawn on the sheets are not, and Excel will recalculate when it opens.")
    wb.save(output_path)
    return report
